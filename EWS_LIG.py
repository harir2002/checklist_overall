import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io


ews1 = []
ews2 = []
ews3 = []
lig1 = []
lig2 = []
lig3 = []




def EWS1(sheet, ignore_year, ignore_month):
    st.write("Analyzing Ews Tower 1")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    ews1.append(1)
                else:
                    ews1.append(0)
            else:
                ews1.append(0)
def EWS2(sheet, ignore_year, ignore_month):
    st.write("Analyzing Ews Tower 2")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['S', 'U', 'W', 'Y', 'AA', 'AC', 'AE', 'AG']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    ews2.append(1)
                else:
                    ews2.append(0)
            else:
                ews2.append(0)

def EWS3(sheet, ignore_year, ignore_month):
    st.write("Analyzing Ews Tower 3")
    rows = [8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22]
    cols = ['AJ', 'AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    ews3.append(1)
                else:
                    ews3.append(0)
            else:
                ews3.append(0)


def LIG1(sheet, ignore_year, ignore_month):
    st.write("Analyzing Lig Tower 1")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['AJ', 'AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    lig1.append(1)
                else:
                    lig1.append(0)
            else:
                lig1.append(0)
    

def LIG2(sheet, ignore_year, ignore_month):
    st.write("Analyzing Lig Tower 2")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['S', 'U', 'W', 'Y', 'AA', 'AC', 'AE', 'AG']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    lig2.append(1)
                else:
                    lig2.append(0)
            else:
                lig2.append(0)

def LIG3(sheet, ignore_year, ignore_month):
    st.write("Analyzing Lig Tower 3")
    rows = [30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    lig3.append(1)
                else:
                    lig3.append(0)
            else:
                lig3.append(0)




def ProcessEWS_LIG(exceldatas, ignore_year, ignore_month):
    wb = load_workbook(exceldatas)
    sheet_name = "Revised Baseline 45daysNGT+Rai"
    sheet = wb[sheet_name]

    ews1.clear()
    EWS1(sheet, ignore_year, ignore_month)
    ews2.clear()
    EWS2(sheet, ignore_year, ignore_month)
    ews3.clear()
    EWS3(sheet, ignore_year, ignore_month)
    lig1.clear()
    LIG1(sheet, ignore_year, ignore_month)
    lig2.clear()
    LIG2(sheet, ignore_year, ignore_month)
    lig3.clear()
    LIG3(sheet, ignore_year, ignore_month)

    data = {
        "Project Name": ["EWS", "EWS", "EWS", "LIG", "LIG", "LIG"],
        "Tower": ["EWST1", "EWST2", "EWST3", "LIGT1", "LIGT2", "LIGT3"],
        "Green (1)": [ews1.count(1), ews2.count(1), ews3.count(1), lig1.count(1), lig2.count(1), lig3.count(1)],
        "Non-Green (0)": [ews1.count(0), ews2.count(0), ews3.count(0), lig1.count(0), lig2.count(0), lig3.count(0)],
       
    }

    project_and_green = [{"Tower": project, "Green (1)": green} for project, green in zip(data["Tower"], data["Green (1)"])]
    json_data = json.dumps(project_and_green, indent=4)


    # st.write(json_data)
    return json_data


    
    

           

       

   

    


    



