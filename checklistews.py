import streamlit as st
import requests
import json
import urllib.parse
import urllib3
import certifi
import pandas as pd
from datetime import datetime
import re
import logging
import os
from dotenv import load_dotenv
import aiohttp
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import openpyxl
import io
from dotenv import load_dotenv
from uuid import uuid4
import ibm_boto3
from ibm_botocore.client import Config
from tenacity import retry, stop_after_attempt, wait_exponential
import xlsxwriter
from EWS_LIG import *
from dateutil.relativedelta import relativedelta
import traceback



# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# IBM COS Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
COS_BUCKET = os.getenv("COS_BUCKET")

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")

# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"


import time
from functools import wraps
import streamlit as st

def function_timer(show_args=False):
    """Decorator to measure and display function execution time"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            # Start timer
            start_time = time.time()
            
            # Call the original function
            result = func(*args, **kwargs)
            
            # Calculate duration
            duration = time.time() - start_time
            
            # Display timing info
            func_name = func.__name__.replace('_', ' ').title()
            arg_info = ""
            if show_args and args:
                arg_info = f" with args: {args[1:]}"  # Skip self if present
            
            st.info(f"⏱️ {func_name}{arg_info} executed in {duration:.2f} seconds")
            
            return result
        return wrapper
    return decorator

if "slabreport" not in st.session_state:
    st.session_state.slabreport = {}



# Login Function
@function_timer()
async def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
    if response.status_code == 200:
        try:
            session_id = response.json().get("UserProfile", {}).get("Sessionid")
            logger.info(f"Login successful, Session ID: {session_id}")
            st.session_state.sessionid = session_id
            st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
            return session_id
        except json.JSONDecodeError:
            logger.error("JSONDecodeError during login")
            st.sidebar.error("❌ Failed to parse login response")
            return None
    logger.error(f"Login failed: {response.status_code} - {response.text}")
    st.sidebar.error(f"❌ Login failed: {response.status_code} - {response.text}")
    return None

# Function to generate access token
@function_timer()
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
def get_access_token(API_KEY):
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
    response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
    try:
        if response.status_code == 200:
            token_info = response.json()
            logger.info("Access token generated successfully")
            return token_info['access_token']
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
            raise Exception("Failed to get access token")
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f"❌ Error getting access token: {str(e)}")
        return None

# Initialize COS client
@function_timer()
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
def initialize_cos_client():
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        logger.info("COS client initialized successfully")
        return cos_client
    except Exception as e:
        logger.error(f"Error initializing COS client: {str(e)}")
        st.error(f"❌ Error initializing COS client: {str(e)}")
        raise

# Fetch Workspace ID
async def GetWorkspaceID():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch workspace list: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch workspace list: {response.status_code}")
    try:
        data = response.json()
        st.session_state.workspaceid = data['asiteDataList']['workspaceVO'][0]['Workspace_Id']
        st.write(f"Workspace ID: {st.session_state.workspaceid}")
    except (KeyError, IndexError) as e:
        st.error(f"Error parsing workspace ID: {str(e)}")
        raise

# Fetch Project IDs
async def GetProjectId():
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch project IDs: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch project IDs: {response.status_code}")
    data = response.json()
    if not data.get('data'):
        st.error("No quality plans found for the specified date.")
        raise Exception("No quality plans found")
    st.session_state.EWS_LIG_structure = data['data'][0]['planId']
    st.write(f"EWS_LIG Structure Project ID: {st.session_state.EWS_LIG_structure}")

# Asynchronous Fetch Function
async def fetch_data(session, url, headers):
    async with session.get(url, headers=headers) as response:
        if response.status == 200:
            return await response.json()
        elif response.status == 204:
            return None
        else:
            raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# Fetch All Structure Data
async def GetAllDatas():
    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_structure_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching EWS_LIG  Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                async with session.get(url, headers=headers) as response:
                    if response.status == 204:
                        st.write("No more EWS_LIG Structure data available (204)")
                        break
                    data = await response.json()
                    if 'associationList' in data and data['associationList']:
                        all_structure_data.extend(data['associationList'])
                    else:
                        all_structure_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_structure_data[-record_limit:])} EWS_LIG Structure records (Total: {len(all_structure_data)})")
                    if len(all_structure_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure data: {str(e)}")
                break

    df_structure = pd.DataFrame(all_structure_data)
    
    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_structure.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_structure.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("❌ Neither statusName nor statusColor found in data!")
        return pd.DataFrame()

    EWS_LIG_structure = df_structure[desired_columns]    

    st.write(f"EWS_LIG STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(EWS_LIG_structure)}")
    st.write(EWS_LIG_structure)  
    
    return EWS_LIG_structure

# Fetch Activity Data
@function_timer()
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_activity_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Activity data for EWS_LIG Structure...")  
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
                break
 
    structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

    st.write("EWS_LIG STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)
      
    return structure_activity_data

# Fetch Location/Module Data
@function_timer()
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_location_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching EWS_LIG Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Location data: {str(e)}")
                break
        
    structure_df = pd.DataFrame(all_structure_location_data)
    
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("❌ All 'name' values in Structure Location data are missing or empty!")

    st.write("EWS_LIG STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)
    
    st.session_state.structure_location_data = structure_df
    
    return structure_df

# Process individual chunk
@function_timer()
def process_chunk(chunk, chunk_idx, dataset_name, location_df):
    logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
    generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
    logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
    return generated_text, chunk_idx

# Process data with manual counting
@function_timer()
def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return "No completed activities found."

    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved EWS_LIG {dataset_name} data to EWS_LIG_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    location_df = st.session_state.structure_location_data

    chunk_results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_chunk = {
            executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx 
            for idx, chunk in enumerate(chunks)
        }

        completed_chunks = 0
        for future in as_completed(future_to_chunk):
            chunk_idx = future_to_chunk[future]
            try:
                generated_text, idx = future.result()
                chunk_results[idx] = generated_text
                completed_chunks += 1
                progress_percent = completed_chunks / len(chunks)
                progress_bar.progress(progress_percent)
                status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
                st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

    parsed_data = {}
    for chunk_idx in sorted(chunk_results.keys()):
        generated_text = chunk_results[chunk_idx]
        if not generated_text:
            continue

        current_tower = None
        tower_activities = []
        lines = generated_text.split("\n")
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            if line.startswith("Tower:"):
                try:
                    tower_parts = line.split("Tower:", 1)
                    if len(tower_parts) > 1:
                        if current_tower and tower_activities:
                            if current_tower not in parsed_data:
                                parsed_data[current_tower] = []
                            parsed_data[current_tower].extend(tower_activities)
                            tower_activities = []
                        current_tower = tower_parts[1].strip()
                except Exception as e:
                    logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
                    if not current_tower:
                        current_tower = f"Unknown Tower {chunk_idx}"
                    
            elif line.startswith("Total Completed Activities:"):
                continue
            elif not line.strip().startswith("activityName"):
                try:
                    parts = re.split(r'\s{2,}', line.strip())
                    if len(parts) >= 2:
                        activity_name = ' '.join(parts[:-1]).strip()
                        count_str = parts[-1].strip()
                        count_match = re.search(r'\d+', count_str)
                        if count_match:
                            count = int(count_match.group())
                            if current_tower:
                                tower_activities.append({
                                    "activityName": activity_name,
                                    "completedCount": count
                                })
                    else:
                        match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
                        if match and current_tower:
                            activity_name = match.group(1).strip()
                            count = int(match.group(2).strip())
                            tower_activities.append({
                                "activityName": activity_name,
                                "completedCount": count
                            })
                except (ValueError, IndexError) as e:
                    logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)

    aggregated_data = {}
    for tower_name, activities in parsed_data.items():
        tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
        if tower_short_name not in aggregated_data:
            aggregated_data[tower_short_name] = {}
        
        for activity in activities:
            name = activity.get("activityName", "Unknown")
            count = activity.get("completedCount", 0)
            if name in aggregated_data[tower_short_name]:
                aggregated_data[tower_short_name][name] += count
            else:
                aggregated_data[tower_short_name][name] = count

    combined_output_lines = []
    sorted_towers = sorted(aggregated_data.keys())
    
    for i, tower_short_name in enumerate(sorted_towers):
        combined_output_lines.append(f"{tower_short_name:<11} activityName            CompletedCount")
        activity_dict = aggregated_data[tower_short_name]
        tower_total = 0
        for name, count in sorted(activity_dict.items()):
            combined_output_lines.append(f"{'':<11} {name:<23} {count:>14}")
            tower_total += count
        combined_output_lines.append(f"{'':<11} Total for {tower_short_name:<11}: {tower_total:>14}")
        if i < len(sorted_towers) - 1:
            combined_output_lines.append("")
    
    combined_output = "\n".join(combined_output_lines)
    return combined_output

# Local formatting function for manual counting
@function_timer()
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
    towers_data = {}
    
    for _, row in chunk.iterrows():
        tower_name = row['tower_name']
        activity_name = row['activityName']
        count = int(row['CompletedCount'])
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        towers_data[tower_name].append({
            "activityName": activity_name,
            "completedCount": count
        })
    
    output = ""
    total_activities = 0
    
    for tower_name, activities in sorted(towers_data.items()):
        output += f"Tower: {tower_name}\n"
        output += "activityName            CompletedCount\n"
        activity_dict = {}
        for activity in activities:
            name = activity['activityName']
            count = activity['completedCount']
            activity_dict[name] = activity_dict.get(name, 0) + count
        for name, count in sorted(activity_dict.items()):
            output += f"{name:<30} {count}\n"
            total_activities += count
    
    output += f"Total Completed Activities: {total_activities}"
    return output

def get_tower_name(full_path):
    parts = full_path.split('/')
    if len(parts) < 2:
        logger.warning(f"Invalid full_path format: {full_path}. Returning as-is.")
        return full_path

    logger.info(f"Processing path: {full_path}")
    logger.info(f"Path parts: {parts}")

    path_lower = full_path.lower()
    is_ews = "ews" in path_lower
    is_lig = "lig" in path_lower

    tower = None
    tower_number = None
    found_part = None
    
    for i, part in enumerate(parts):
        logger.info(f"Checking part {i}: '{part}'")
        
        # Look for "Tower X" or "Pour X" format
        part_lower = part.lower()
        if part_lower.startswith("tower"):
            tower = part
            found_part = f"Tower format: {part}"
            tower_parts = part.split()
            if len(tower_parts) > 1 and tower_parts[1].isdigit():
                tower_number = tower_parts[1]
            else:
                logger.warning(f"Could not extract tower number from: {part}")
                tower_number = "Unknown"
            break
        elif part_lower.startswith("pour"):
            found_part = f"Pour format: {part}"
            pour_parts = part.split()
            if len(pour_parts) > 1 and pour_parts[1].isdigit():
                tower_number = pour_parts[1]
                tower = f"Pour {tower_number}"
            else:
                logger.warning(f"Could not extract pour number from: {part}")
                tower_number = "Unknown"
                tower = part
            break
        elif "tower" in part_lower or "pour" in part_lower:
            found_part = f"Contains 'tower' or 'pour': {part}"
            logger.info(f"Found part containing 'tower' or 'pour': {part}")
            number_match = re.search(r'\d+', part)
            if number_match:
                tower_number = number_match.group()
                tower = f"Tower {tower_number}" if "tower" in part_lower else f"Pour {tower_number}"
            else:
                logger.warning(f"Could not extract number from: {part}")
                tower_number = "Unknown"
                tower = part
            break
    
    logger.info(f"Found part: {found_part}")
    logger.info(f"Tower: {tower}, Tower number: {tower_number}")
    
    if not tower:
        logger.warning(f"Tower/Pour name not found in path: {full_path}. Returning as-is.")
        return full_path

    if is_ews and not is_lig:
        prefix = "EWS"
    elif is_lig and not is_ews:
        prefix = "LIG"
    elif is_ews and is_lig:
        # Handle cases where both EWS and LIG appear in the path
        ews_idx = path_lower.find("ews")
        lig_idx = path_lower.find("lig")
        prefix = "EWS" if ews_idx < lig_idx else "LIG"
        logger.info(f"Both EWS and LIG in path, using prefix: {prefix}")
    else:
        logger.warning(f"Could not classify EWS/LIG for path: {full_path}. Defaulting to 'Unknown' prefix.")
        prefix = "Unknown"

    tower_name = f"{prefix} {tower}"
    logger.info(f"Final tower name: {tower_name}")
    return tower_name

def get_full_path(location_id, parent_child_dict, name_dict):
    path = []
    current_id = location_id
    max_depth = 15
    depth = 0
    visited_ids = set()
    
    while current_id and depth < max_depth:
        if current_id in visited_ids:
            logger.warning(f"Circular reference detected for location_id {location_id} at {current_id}. Path so far: {path}")
            break
        visited_ids.add(current_id)
        
        if current_id not in parent_child_dict or current_id not in name_dict:
            logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
            break
        
        parent_id = parent_child_dict.get(current_id)
        name = name_dict.get(current_id, "Unknown")
        path.append(name)
        
        if not parent_id:
            break
        
        current_id = parent_id
        depth += 1
    
    if depth >= max_depth:
        logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible deep hierarchy or error. Path: {path}")
    
    if not path:
        logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
        return "Unknown"
    
    full_path = '/'.join(reversed(path))
    logger.debug(f"Full path for location_id {location_id}: {full_path}")
    return full_path

@function_timer()
def is_roof_slab_only(full_path):
    parts = full_path.split('/')
    last_part = parts[-1].lower()
    is_slab = any(keyword in last_part for keyword in ['roof slab', 'slab', 'roofslab', 'slab level'])
    logger.debug(f"Checking roof slab for path: {full_path}, result: {is_slab}")
    return is_slab


@function_timer()
def process_data(df, activity_df, location_df, dataset_name, use_module_hierarchy_for_finishing=False):
    completed = df[df['statusName'] == 'Completed'].copy()
    
    asite_activities = [
        "Wall Conducting", "Plumbing Works", "POP & Gypsum Plaster", "Wiring & Switch Socket",
        "Slab Conducting", "Electrical Cable", "Door/Window Frame", "Waterproofing - Sunken",
        "Wall Tile", "Floor Tile", "Door/Window Shutter", "Shuttering", "Reinforcement",
        "Sewer Line", "Rain Water/Storm Line", "Granular Sub-base", "WMM",
        "Saucer drain/Paver block", "Kerb Stone", "Concreting"
    ]
    
    count_table = pd.DataFrame({'Count': [0] * len(asite_activities)}, index=asite_activities)
    
    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data.")
        return pd.DataFrame(), 0, count_table

    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

    if 'qiActivityId' not in completed.columns:
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

    if completed['name'].isna().all():
        logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
        st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
        completed['name'] = 'Unknown'
    else:
        completed['name'] = completed['name'].fillna('Unknown')

    def normalize_activity_name(name):
        typo_corrections = {
            "Wall Conduting": "Wall Conducting",
            "Slab conduting": "Slab Conducting",
            "WallTile": "Wall Tile",
            "FloorTile": "Floor Tile",
            "wall tile": "Wall Tile",
            "floor tile": "Floor Tile",
            "DoorWindowFrame": "Door/Window Frame",
            "DoorWindowShutter": "Door/Window Shutter",
            "Second Roof Slab": "Roof Slab",
            "First Roof Slab": "Roof Slab",
            "Roof slab": "Roof Slab",
            "Beam": "Beam",
            "Column": "Column",
            "Reinforcement": "Reinforcement",
            "Shuttering": "Shuttering",
            "Concreting": "Concreting",
            "DeShuttering": "De-Shuttering"
        }
        for typo, correct in typo_corrections.items():
            if name.lower() == typo.lower():
                return correct
        return name

    completed['activityName'] = completed['activityName'].apply(normalize_activity_name).fillna('Unknown')

    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    completed['full_path'] = completed['qiLocationId'].apply(
        lambda x: get_full_path(x, parent_child_dict, name_dict)
    )

    logger.debug(f"All unique full_path values in {dataset_name} dataset BEFORE filtering:")
    full_path_counts = completed['full_path'].value_counts()
    for path, count in full_path_counts.items():
        logger.debug(f"  Path: {path}, Count: {count}")

    completed['temp_tower_name'] = completed['full_path'].apply(
        lambda x: x.split('/')[1] if len(x.split('/')) > 1 and ('Tower' in x.split('/')[1] or 'Pour' in x.split('/')[1]) else 'Unknown'
    )
    tower_counts_before = completed['temp_tower_name'].value_counts()
    logger.debug(f"Tower distribution BEFORE filtering in {dataset_name}:")
    for tower, count in tower_counts_before.items():
        logger.debug(f"  {tower}: {count} records")

    def has_flat_number(full_path):
        parts = full_path.split('/')
        last_part = parts[-1]
        match = re.match(r'^\d+(?:(?:\s*\(LL\))|(?:\s*\(UL\))|(?:\s*LL)|(?:\s*UL))?$', last_part)
        return bool(match)

    def is_roof_slab_only(full_path):
        parts = full_path.split('/')
        last_part = parts[-1].lower()
        is_slab = any(keyword in last_part for keyword in ['roof slab', 'slab', 'roofslab', 'slab level'])
        logger.debug(f"Checking roof slab for path: {full_path}, result: {is_slab}")
        return is_slab

    if dataset_name.lower() == 'structure':
        logger.debug(f"Applying roof slab filtering for {dataset_name} dataset")
        completed_before_filter = len(completed)
        
        logger.debug(f"All unique paths before roof slab filtering:")
        for path, count in full_path_counts.items():
            logger.debug(f"  Path: {path}, Count: {count}")
        
        # Log paths that will be filtered out
        logger.debug("Paths that WILL be filtered out by is_roof_slab_only:")
        paths_to_be_filtered = completed[~completed['full_path'].apply(is_roof_slab_only)]['full_path'].unique()
        for path in sorted(paths_to_be_filtered):
            logger.debug(f"  ✗ Path: {path}")
        
        completed = completed[completed['full_path'].apply(is_roof_slab_only)]
        completed_after_filter = len(completed)
        logger.debug(f"Roof slab filtering: {completed_before_filter} -> {completed_after_filter} records")
        
        if not completed.empty:
            logger.debug(f"Paths that passed roof slab filtering:")
            full_path_counts_after = completed['full_path'].value_counts()
            for path, count in full_path_counts_after.items():
                logger.debug(f"  ✓ Path: {path}, Count: {count}")
        else:
            logger.warning(f"No paths contain 'roof slab', 'slab', 'roofslab', or 'slab level' in {dataset_name} dataset")
    
    else:
        completed = completed[completed['full_path'].apply(has_flat_number)]
        if completed.empty:
            logger.warning(f"No completed activities with flat numbers found in {dataset_name} data after filtering.")
            return pd.DataFrame(), 0, count_table

    completed['temp_tower_name'] = completed['full_path'].apply(
        lambda x: x.split('/')[1] if len(x.split('/')) > 1 and ('Tower' in x.split('/')[1] or 'Pour' in x.split('/')[1]) else 'Unknown'
    )
    tower_counts_after = completed['temp_tower_name'].value_counts()
    logger.debug(f"Tower distribution AFTER filtering in {dataset_name}:")
    for tower, count in tower_counts_after.items():
        logger.debug(f"  {tower}: {count} records")
    completed = completed.drop(columns=['temp_tower_name'])

    completed['tower_name'] = completed['full_path'].apply(get_tower_name)

    logger.debug(f"All tower_name values after get_tower_name in {dataset_name}:")
    tower_name_counts = completed['tower_name'].value_counts()
    for tower_name, count in tower_name_counts.items():
        logger.debug(f"  {tower_name}: {count} records")

    logger.debug(f"Sample full_path to tower_name mapping in {dataset_name}:")
    for idx, row in completed[['full_path', 'tower_name']].head(20).iterrows():
        logger.debug(f"  full_path: {row['full_path']} -> tower_name: {row['tower_name']}")

    analysis = completed.groupby(['tower_name', 'activityName'])['qiLocationId'].nunique().reset_index(name='CompletedCount')
    analysis = analysis.sort_values(by=['tower_name', 'activityName'], ascending=True)
    total_completed = analysis['CompletedCount'].sum()

    activity_counts = completed.groupby('activityName')['qiLocationId'].nunique().reset_index(name='Count')
    for activity in asite_activities:
        if activity in activity_counts['activityName'].values:
            count_table.loc[activity, 'Count'] = activity_counts[activity_counts['activityName'] == activity]['Count'].iloc[0]

    logger.info(f"Total completed activities for {dataset_name}: {total_completed}")
    logger.info(f"Count table for {dataset_name}:\n{count_table.to_string()}")
    
    logger.debug(f"Final analysis results for {dataset_name} by tower:")
    for tower in sorted(analysis['tower_name'].unique()):
        tower_data = analysis[analysis['tower_name'] == tower]
        tower_total = tower_data['CompletedCount'].sum()
        logger.debug(f"  {tower}: {tower_total} total completed activities")
    
    return analysis, total_completed, count_table


# Main analysis function for Wave City Club Structure
@function_timer()
def AnalyzeStatusManually(email=None, password=None):
    start_time = time.time()

    if 'sessionid' not in st.session_state:
        st.error("❌ Please log in first!")
        logger.error("AnalyzeStatusManually failed: No sessionid in st.session_state")
        return

    required_data = [
        'eden_structure',
        'structure_activity_data',
        'structure_location_data'
    ]
    
    for data_key in required_data:
        if data_key not in st.session_state:
            st.error(f"❌ Please fetch required data first! Missing: {data_key}")
            logger.error(f"AnalyzeStatusManually failed: Missing {data_key} in st.session_state")
            return
        if not isinstance(st.session_state[data_key], pd.DataFrame):
            st.error(f"❌ {data_key} is not a DataFrame! Found type: {type(st.session_state[data_key])}")
            logger.error(f"AnalyzeStatusManually failed: {data_key} is not a DataFrame, found type {type(st.session_state[data_key])}")
            return
        if st.session_state[data_key].empty:
            st.error(f"❌ {data_key} is an empty DataFrame!")
            logger.error(f"AnalyzeStatusManually failed: {data_key} is an empty DataFrame")
            return

    structure_data = st.session_state.eden_structure
    structure_activity = st.session_state.structure_activity_data
    structure_locations = st.session_state.structure_location_data
    
    for df, name in [(structure_data, "Structure")]:
        required_columns = ['statusName', 'qiLocationId', 'activitySeq']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns {missing_columns} in {name} data!")
            logger.error(f"AnalyzeStatusManually failed: Missing columns {missing_columns} in {name} data")
            return

    for df, name in [(structure_locations, "Structure Location")]:
        required_columns = ['qiLocationId', 'name']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns {missing_columns} in {name} data!")
            logger.error(f"AnalyzeStatusManually failed: Missing columns {missing_columns} in {name} data")
            return

    for df, name in [(structure_activity, "Structure Activity")]:
        required_columns = ['activitySeq', 'activityName']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns {missing_columns} in {name} data!")
            logger.error(f"AnalyzeStatusManually failed: Missing columns {missing_columns} in {name} data")
            return

    try:
        logger.info("Starting structure data processing...")
        structure_analysis, structure_total, _ = process_data(structure_data, structure_activity, structure_locations, "Structure")
        logger.info(f"Structure data processed. Total completed activities: {structure_total}")
    except Exception as e:
        st.error(f"❌ Error processing structure data: {str(e)}")
        logger.error(f"AnalyzeStatusManually failed: Error processing structure data: {str(e)}")
        logger.error(f"Stack trace:\n{traceback.format_exc()}")
        return

    st.session_state.structure_analysis = structure_analysis
    st.session_state.structure_total = structure_total
    logger.info("Structure analysis stored in st.session_state")

    if 'tower_name' not in structure_analysis.columns:
        st.error("❌ Tower names not found in structure analysis. Check location data and tower extraction logic.")
        logger.error("AnalyzeStatusManually failed: tower_name column not found in structure_analysis")
        return

    unique_towers = structure_analysis['tower_name'].unique()
    logger.debug(f"Unique towers in structure_analysis: {list(unique_towers)}")

    if len(unique_towers) <= 1:
        logger.warning(f"Only {len(unique_towers)} tower(s) found: {unique_towers}. Expected multiple towers for EWS LIG.")
        st.warning(f"Only {len(unique_towers)} tower(s) found. This may indicate an issue with location data separation.")

    def sort_key(tower_name):
        prefix = tower_name.split()[0]  # EWS or LIG
        tower_num = int(tower_name.split()[-1]) if tower_name.split()[-1].isdigit() else 0
        return (prefix, tower_num)

    sorted_towers = sorted(unique_towers, key=sort_key)

    st.write("### EWS_LIG Structure Quality Analysis (Completed Activities):")
    
    for tower in sorted_towers:
        tower_data = structure_analysis[structure_analysis['tower_name'] == tower]
        if tower_data.empty:
            st.write(f"**{tower}:** No completed activities found.")
            logger.info(f"No completed activities for {tower}")
            continue

        st.write(f"**{tower}:**")
        output_lines = []
        output_lines.append("activityName            CompletedCount")
        tower_total = 0
        for _, row in tower_data.iterrows():
            output_lines.append(f"{row['activityName']:<30} {row['CompletedCount']}")
            tower_total += row['CompletedCount']
        output_lines.append(f"{'Total for ' + tower:<30} {tower_total}")
        st.text("\n".join(output_lines))
        logger.info(f"Displayed output for {tower}")

    st.write(f"**Total Completed Activities Across All Towers:** {structure_total}")

    end_time = time.time()
    st.write(f"Total execution time: {end_time - start_time:.2f} seconds")
    logger.info(f"AnalyzeStatusManually completed in {end_time - start_time:.2f} seconds")



@function_timer()
def get_cos_files():
    try:
        # Initialize COS client
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client. Check credentials or configuration.")
            logger.error("Failed to initialize COS client")
            return None

        # Step 1: List all objects in the bucket to inspect structure
        st.write(f"Listing all objects in bucket '{COS_BUCKET}' (no prefix)")
        response = cos_client.list_objects_v2(Bucket=COS_BUCKET)
        if 'Contents' not in response:
            st.error(f"❌ No objects found in bucket '{COS_BUCKET}'. Verify bucket name and permissions.")
            logger.error(f"No objects found in bucket {COS_BUCKET}")
            return None

        all_files = [obj['Key'] for obj in response.get('Contents', [])]
        st.write("**All files in bucket:**")
        if all_files:
            st.write("\n".join(all_files))
        else:
            st.write("No files found in the bucket.")
            logger.warning(f"Bucket {COS_BUCKET} is empty")
            return None

        # Extract folder names (prefixes)
        folders = set()
        for file in all_files:
            if '/' in file:
                folder = file.split('/')[0] + '/'
                folders.add(folder)
        st.write("**Available folders in bucket:**")
        st.write("\n".join(folders) if folders else "No folders found.")

        # Step 2: Focus on the EWS LIG P4 folder with variations
        possible_prefixes = [
            "EWS LIG P4/",  # Exact match
            "EWS LIG P4",   # Without trailing slash
            "ews lig p4/",  # Lowercase
            "EWS LIG P4 /", # Extra space
            "EWS_LIG_P4/",  # Underscores instead of spaces
            "EWS-LIG-P4/",  # Hyphens instead of spaces
        ]

        target_files = []
        for prefix in possible_prefixes:
            st.write(f"\nListing objects in bucket '{COS_BUCKET}' with prefix '{prefix}'")
            response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix=prefix)
            
            if 'Contents' not in response:
                st.write(f"No files found in '{prefix}' folder.")
                logger.info(f"No objects found in {prefix} folder")
                continue

            prefix_files = [obj['Key'] for obj in response.get('Contents', [])]
            st.write(f"**Files in {prefix} folder:**")
            if prefix_files:
                st.write("\n".join(prefix_files))
            else:
                st.write("No files found.")
                logger.info(f"{prefix} folder is empty")
                continue

            # Updated regex pattern to match "Structure Work Tracker" instead of "Checklist Report"
            pattern = re.compile(
                r"(?i)EWS\s*LIG\s*P4/.*?Structure\s*Work\s*Tracker.*?[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$"
            )

            # Supported date formats for parsing
            date_formats = [
                "%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y",
                "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d",
                "%d%m%Y", "%Y.%m.%d"
            ]

            for key in prefix_files:
                match = pattern.match(key)
                if match:
                    date_str = match.group(1).strip('()').strip()
                    parsed_date = None
                    for fmt in date_formats:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt)
                            break
                        except ValueError:
                            continue
                    if parsed_date:
                        target_files.append({'key': key, 'date': parsed_date})
                    else:
                        logger.warning(f"Could not parse date in filename: {key}")
                        st.warning(f"Skipping file with unparseable date: {key}")
                else:
                    st.write(f"File '{key}' does not match the expected pattern.")

        if not target_files:
            st.error(f"❌ No Excel files matched the expected pattern in any of the folders: {', '.join(possible_prefixes)}")
            logger.error("No files matched the expected pattern")
            return None

        # Find the latest file based on the parsed date
        latest_file = max(target_files, key=lambda x: x['date'])
        file_key = latest_file['key']
        st.success(f"Found matching file: {file_key}")
        return file_key

    except Exception as e:
        st.error(f"❌ Error fetching COS files: {str(e)}")
        logger.error(f"Error fetching COS files: {str(e)}")
        return None

if 'cos_df_Revised_Baseline_45daysNGT_Rai' not in st.session_state:
    st.session_state.cos_df_Revised_Baseline_45daysNGT_Rai = None

if 'ai_response' not in st.session_state:
    st.session_state.ai_response = {} 

# Process Excel files for Wave City Club blocks with updated sheet names and expected_columns
@function_timer()
def process_file(file_stream, filename):
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")

        target_sheets = ["Revised Baseline 45daysNGT+Rai"]
        results = []

        for sheet_name in target_sheets:
            if sheet_name not in available_sheets:
                st.warning(f"Sheet '{sheet_name}' not found in file: {filename}")
                continue

            file_stream.seek(0)

            try:
                # Read the first few rows to inspect the data
                df_preview = pd.read_excel(file_stream, sheet_name=sheet_name, nrows=10)
                st.write(f"Preview of first 10 rows in {sheet_name}:")
                st.write(df_preview)

                # Try different header rows
                header_found = False
                actual_finish_col = None
                for header_row in [4, 5, 6, 3, 2]:
                    file_stream.seek(0)
                    df = pd.read_excel(file_stream, sheet_name=sheet_name, header=header_row)
                    st.write(f"Testing header row {header_row} in {sheet_name}. Raw columns: {list(df.columns)}")

                    df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]

                    # Check for 'Floors' or floor identifiers
                    if 'Floors' in df.columns or any('Floor' in str(col) for col in df.columns):
                        header_found = True
                    elif not df.empty and any(str(df.iloc[i, 0]).strip() in ['GF', '1F', '2F', '3F', '4F', '5F'] for i in range(min(5, len(df)))):
                        if df.columns[0] != 'Floors':
                            df.rename(columns={df.columns[0]: 'Floors'}, inplace=True)
                        header_found = True

                    # Check for 'Actual Finish' or variants
                    for col in df.columns:
                        if str(col).lower() in ['actual finish', 'actual_finish', 'finish date', 'completion date']:
                            actual_finish_col = col
                            break

                    if header_found and actual_finish_col:
                        break

                if not header_found:
                    st.error(f"No valid header row found in {sheet_name}. Expected to find 'Floors' column or floor identifiers.")
                    continue

                # Clean up the dataframe
                df = df.dropna(subset=[df.columns[0]])
                df = df[~df.iloc[:, 0].astype(str).str.contains('Floor|Pour|Baseline|Days', case=False, na=False)]
                
                floor_pattern = r'^(GF|\d{1,2}F)$'
                df = df[df.iloc[:, 0].astype(str).str.match(floor_pattern, na=False)]

                df.rename(columns={df.columns[0]: 'Activity Name'}, inplace=True)

                # Rename 'Actual Finish' if found
                if actual_finish_col:
                    df.rename(columns={actual_finish_col: 'Actual Finish'}, inplace=True)
                else:
                    st.warning(f"No 'Actual Finish' column found in {sheet_name}. Adding empty column.")
                    df['Actual Finish'] = pd.NA
                    logger.warning(f"No 'Actual Finish' column in {sheet_name}")

                target_columns = ['Activity Name', 'Actual Finish']
                available_columns = [col for col in target_columns if col in df.columns]
                for col in df.columns:
                    if col not in target_columns:
                        available_columns.append(col)

                if len(available_columns) <= 1:
                    st.error(f"Only 'Activity Name' found in {sheet_name}. No additional columns to process.")
                    continue

                df = df[available_columns]
                df['Activity Name'] = df['Activity Name'].astype(str).str.strip()

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    st.write(f"Sample of rows with NA or invalid values in Actual Finish for {sheet_name}:")
                    na_rows = df[has_na_mask][['Activity Name', 'Actual Finish']]
                    if not na_rows.empty:
                        st.write(na_rows.head(10))
                    else:
                        st.write("No NA or invalid values found in Actual Finish")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                st.write(f"Unique Activity Names (Floor identifiers) in {sheet_name}:")
                st.write(df[['Activity Name']].drop_duplicates())

                st.write(f"Final processed dataframe shape: {df.shape}")
                st.write(f"Final columns: {list(df.columns)}")
                st.write("Sample of processed data:")
                st.write(df.head())

                results.append((df, sheet_name))

            except Exception as e:
                st.error(f"Error processing sheet {sheet_name}: {str(e)}")
                logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                continue

        if not results:
            st.error(f"No valid sheets ({', '.join(target_sheets)}) found in file: {filename}")
            return [(None, None)]

        return results

    except Exception as e:
        st.error(f"Error loading Excel file: {str(e)}")
        logger.error(f"Error loading Excel file: {str(e)}")
        return [(None, None)]

# Function to get access token for WatsonX API
@function_timer()
def get_access_token(api_key):
    try:
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {
            "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
            "apikey": api_key
        }
        response = requests.post("https://iam.cloud.ibm.com/identity/token", headers=headers, data=data)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        logger.error(f"Error getting access token: {str(e)}")
        return None

#Slab code
@function_timer()
def GetSlabReport():
    st.write("EWS LIG Structure Work Tracker")
    found_verdia = False
    today = datetime.today()
    current_month_year = today.strftime("%m-%Y")  # Current month and year
    prev_month = today - relativedelta(months=1)
    prev_month_year = prev_month.strftime("%m-%Y")  # Previous month and year
    
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        
        # List objects in the S3 bucket
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]

        # Try to find the current month's file
        for file in files:
            try:
                if file.startswith("EWS LIG") and "Structure Work Tracker" in file and current_month_year in file:
                    st.write(f"Found current month file: {file}")
                    response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                    st.write(f"Processing file: {file}")
                    
                    if st.session_state.ignore_month and st.session_state.ignore_year:
                        st.session_state.slabreport = ProcessEWS_LIG(io.BytesIO(response['Body'].read()), st.session_state.ignore_year, st.session_state.ignore_month)
                    else:
                        st.session_state.slabreport = ProcessEWS_LIG(io.BytesIO(response['Body'].read()), st.session_state.ignore_year, st.session_state.ignore_month)               
                    
                    found_verdia = True
                    break

            except Exception as e:
                st.error(f"Error processing file {file}: {e}")

        if not found_verdia:
            # Current month's file not found, display a message
            st.warning(f"Current month file ({current_month_year}) not found.")
            st.session_state.slabreport = "Current month file not found."

            # Optionally, try to fetch the previous month's file if needed
            for file in files:
                try:
                    if file.startswith("EWS LIG") and "Structure Work Tracker" in file and prev_month_year in file:
                        st.write(f"Found previous month file: {file}")
                        response = cos_client.get_object(Bucket="projectreportnew", Key=file)
                        st.write(f"Processing previous month file: {file}")

                        if st.session_state.ignore_month and st.session_state.ignore_year:
                            st.session_state.slabreport = ProcessEWS_LIG(io.BytesIO(response['Body'].read()), st.session_state.ignore_year, st.session_state.ignore_month)  
                        else:
                            st.session_state.slabreport = ProcessEWS_LIG(io.BytesIO(response['Body'].read()), st.session_state.ignore_year, st.session_state.ignore_month)
                        break

                except Exception as e:
                    st.error(f"Error processing previous month file {file}: {e}")
                    st.session_state.slabreport = "No Data Found"

    except Exception as e:
        st.error(f"Error fetching COS files: {e}")
        st.session_state.slabreport = "No Data Found"

# WatsonX Prompt Generation
@function_timer()
def generatePrompt(json_datas, tower_name):
    try:
        GetSlabReport()
        st.write(st.session_state.slabreport)
        st.write(json.loads(st.session_state.slabreport))
        if isinstance(json_datas, pd.DataFrame):
            json_str = json_datas.reset_index().to_json(orient='records', indent=2)
        else:
            json_str = str(json_datas)

        body = {
            "input": f"""
            Read the table data provided below for the tower '{tower_name}' and categorize the activities into the following categories: MEP, Interior Finishing, Structure Work, and ED Civil. Compute the total count of each activity within its respective category and return the results as a JSON array, following the example format provided. For the MEP category, calculate the minimum count between 'UP-First Fix' and 'CP-First Fix' and report it as 'Min. count of UP-First Fix and CP-First Fix'. If an activity is not found in the data, include it with a count of 0. If the Structure Work category has no activities, include it as an empty array. Ensure the counts are accurate, the output is grouped by category, and the JSON structure is valid with no nested or repeated keys.
            
            The table data is the data from the EWS_LIG file for '{tower_name}'.
            {st.session_state.slabreport }

            Table Data:
            {json_str}

            Categories and Activities:
            - MEP: EL-First Fix, Min. count of UP-First Fix and CP-First Fix, C-Gypsum and POP Punning, EL-Second Fix, Concreting, Electrical
            - Interior Finishing: Installation of doors, Waterproofing Works, Wall Tiling, Floor Tiling
            - ED Civil: Sewer Line, Storm Line, GSB, WMM, Stamp Concrete, Saucer drain, Kerb Stone
            - Structure Work: (no activities specified)

            Example JSON format needed:
            [
              {{
                "Category": "MEP",
                "Activities": [
                  {{"Activity Name": "EL-First Fix", "Total": 0}},
                  {{"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0}},
                  {{"Activity Name": "C-Gypsum and POP Punning", "Total": 0}},
                  {{"Activity Name": "EL-Second Fix", "Total": 0}},
                  {{"Activity Name": "Concreting", "Total": 0}},
                  {{"Activity Name": "Electrical", "Total": 0}}
                ]
              }},
              {{
                "Category": "Interior Finishing",
                "Activities": [
                  {{"Activity Name": "Installation of doors", "Total": 0}},
                  {{"Activity Name": "Waterproofing Works", "Total": 0}},
                  {{"Activity Name": "Wall Tiling", "Total": 0}},
                  {{"Activity Name": "Floor Tiling", "Total": 0}}
                ]
              }},
              {{
                "Category": "Structure Work",
                "Activities": []
              }},
              {{
                "Category": "ED Civil",
                "Activities": [
                  {{"Activity Name": "Sewer Line", "Total": 0}},
                  {{"Activity Name": "Storm Line", "Total": 0}},
                  {{"Activity Name": "GSB", "Total": 0}},
                  {{"Activity Name": "WMM", "Total": 0}},
                  {{"Activity Name": "Stamp Concrete", "Total": 0}},
                  {{"Activity Name": "Saucer drain", "Total": 0}},
                  {{"Activity Name": "Kerb Stone", "Total": 0}}
                ]
              }}
            {{"Slab":{{"Tower Name":"Total"}}}}
            ]

            Return only the JSON array, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
            """,
            "parameters": {
                "decoding_method": "greedy",
                "max_new_tokens": 8100,
                "min_new_tokens": 0,
                "stop_sequences": [";"],
                "repetition_penalty": 1.05,
                "temperature": 0.5
            },
            "model_id": os.getenv("MODEL_ID_1"),
            "project_id": os.getenv("PROJECT_ID_1")
        }
        
        access_token = get_access_token(os.getenv("API_KEY_1"))
        if not access_token:
            logger.error("Failed to obtain access token for WatsonX API")
            return generate_fallback_totals(json_datas)
            
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        logger.info(f"Sending request to WatsonX API for tower {tower_name}")
        response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=60)
        
        logger.info(f"WatsonX API response status for {tower_name}: {response.status_code}")
        logger.debug(f"WatsonX API response text: {response.text[:1000]}...")
        
        if response.status_code != 200:
            logger.error(f"WatsonX API call failed for {tower_name}: {response.status_code} - {response.text}")
            st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)
            
        response_data = response.json()
        logger.debug(f"WatsonX API response data for {tower_name}: {response_data}")
        
        if 'results' not in response_data or not response_data['results']:
            logger.error(f"WatsonX API response does not contain 'results' key for {tower_name}")
            st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        generated_text = response_data['results'][0].get('generated_text', '').strip()
        logger.debug(f"Generated text from WatsonX for {tower_name}: {generated_text[:1000]}...")
        
        if not generated_text:
            logger.error(f"WatsonX API returned empty generated text for {tower_name}")
            st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        if not (generated_text.startswith('[') and generated_text.endswith(']')):
            start_idx = generated_text.find('[')
            end_idx = generated_text.rfind(']')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                generated_text = generated_text[start_idx:end_idx+1]
                logger.info(f"Extracted JSON array from response for {tower_name}")
            else:
                logger.error(f"Could not extract valid JSON array from response for {tower_name}: {generated_text[:1000]}...")
                return generate_fallback_totals(json_datas)
        
        try:
            parsed_json = json.loads(generated_text)
            if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in parsed_json):
                logger.warning(f"JSON structure doesn't match expected format for {tower_name}")
                return generate_fallback_totals(json_datas)
            logger.info(f"Successfully parsed WatsonX API response for {tower_name}")
            return generated_text
        except json.JSONDecodeError as e:
            logger.error(f"WatsonX API returned invalid JSON for {tower_name}: {e}")
            st.warning(f"WatsonX API returned invalid JSON. Error: {str(e)}. Using fallback method to calculate totals.")
            error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
            context_start = max(0, error_position - 50)
            context_end = min(len(generated_text), error_position + 50)
            logger.error(f"JSON error context: ...{generated_text[context_start:error_position]}[ERROR HERE]{generated_text[error_position:context_end]}...")
            return generate_fallback_totals(json_datas)
    
    except Exception as e:
        logger.error(f"Error in WatsonX API call for {tower_name}: {str(e)}")
        st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
        return generate_fallback_totals(json_datas)

# Fallback Total Calculation
@function_timer()
def generate_fallback_totals(count_table):
    try:
        if not isinstance(count_table, pd.DataFrame):
            logger.error("Fallback method received invalid input: not a DataFrame")
            return json.dumps([
                {"Category": "MEP", "Activities": [
                    {"Activity Name": "EL-First Fix", "Total": 0},
                    {"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0},
                    {"Activity Name": "C-Gypsum and POP Punning", "Total": 0},
                    {"Activity Name": "EL-Second Fix", "Total": 0},
                    {"Activity Name": "Concreting", "Total": 0},
                    {"Activity Name": "Electrical", "Total": 0}
                ]},
                {"Category": "Interior Finishing", "Activities": [
                    {"Activity Name": "Installation of doors", "Total": 0},
                    {"Activity Name": "Waterproofing Works", "Total": 0},
                    {"Activity Name": "Wall Tiling", "Total": 0},
                    {"Activity Name": "Floor Tiling", "Total": 0}
                ]},
                {"Category": "Structure Work", "Activities": []},
                {"Category": "ED Civil", "Activities": [
                    {"Activity Name": "Sewer Line", "Total": 0},
                    {"Activity Name": "Storm Line", "Total": 0},
                    {"Activity Name": "GSB", "Total": 0},
                    {"Activity Name": "WMM", "Total": 0},
                    {"Activity Name": "Stamp Concrete", "Total": 0},
                    {"Activity Name": "Saucer drain", "Total": 0},
                    {"Activity Name": "Kerb Stone", "Total": 0}
                ]}
            ], indent=2)

        categories = {
            "MEP": [
                "EL-First Fix", "Min. count of UP-First Fix and CP-First Fix",
                "C-Gypsum and POP Punning", "EL-Second Fix", "Concreting", "Electrical"
            ],
            "Interior Finishing": [
                "Installation of doors", "Waterproofing Works", "Wall Tiling", "Floor Tiling"
            ],
            "Structure Work": [],
            "ED Civil": [
                "Sewer Line", "Storm Line", "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone"
            ]
        }

        result = []
        for category, activities in categories.items():
            category_data = {"Category": category, "Activities": []}
            
            if category == "MEP":
                for activity in activities:
                    if activity == "Min. count of UP-First Fix and CP-First Fix":
                        combined_count = count_table.loc["UP-First Fix and CP-First Fix", "Count"] if "UP-First Fix and CP-First Fix" in count_table.index else 0
                        total = combined_count
                    else:
                        total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                    category_data["Activities"].append({
                        "Activity Name": activity,
                        "Total": int(total) if pd.notna(total) else 0
                    })
            elif category == "Structure Work":
                category_data["Activities"] = []
            else:
                for activity in activities:
                    total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                    category_data["Activities"].append({
                        "Activity Name": activity,
                        "Total": int(total) if pd.notna(total) else 0
                    })
            
            result.append(category_data)

        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error in fallback total calculation: {str(e)}")
        st.error(f"Error in fallback total calculation: {str(e)}")
        return json.dumps([
            {"Category": "MEP", "Activities": [
                {"Activity Name": "EL-First Fix", "Total": 0},
                {"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0},
                {"Activity Name": "C-Gypsum and POP Punning", "Total": 0},
                {"Activity Name": "EL-Second Fix", "Total": 0},
                {"Activity Name": "Concreting", "Total": 0},
                {"Activity Name": "Electrical", "Total": 0}
            ]},
            {"Category": "Interior Finishing", "Activities": [
                {"Activity Name": "Installation of doors", "Total": 0},
                {"Activity Name": "Waterproofing Works", "Total": 0},
                {"Activity Name": "Wall Tiling", "Total": 0},
                {"Activity Name": "Floor Tiling", "Total": 0}
            ]},
            {"Category": "Structure Work", "Activities": []},
            {"Category": "ED Civil", "Activities": [
                {"Activity Name": "Sewer Line", "Total": 0},
                {"Activity Name": "Storm Line", "Total": 0},
                {"Activity Name": "GSB", "Total": 0},
                {"Activity Name": "WMM", "Total": 0},
                {"Activity Name": "Stamp Concrete", "Total": 0},
                {"Activity Name": "Saucer drain", "Total": 0},
                {"Activity Name": "Kerb Stone", "Total": 0}
            ]}
        ], indent=2)
@function_timer()
def getTotal(ai_data):
    st.write(ai_data)
    try:
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
            
        if not isinstance(ai_data, list):
            logger.error(f"AI data is not a list: {ai_data}")
            return {}

        totals = {}
        for category_data in ai_data:
            if isinstance(category_data, dict) and 'Activities' in category_data:
                for activity in category_data['Activities']:
                    if isinstance(activity, dict) and 'Total' in activity:
                        activity_name = activity['Activity Name']
                        total = activity['Total']
                        totals[activity_name] = int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0
                    else:
                        logger.warning(f"Invalid activity format: {activity}")
            else:
                logger.warning(f"Invalid category format: {category_data}")
        return totals
    except Exception as e:
        logger.error(f"Error parsing AI data: {str(e)}")
        st.error(f"Error parsing AI data: {str(e)}")
        return {}

# Function to handle activity count display
@function_timer()
def display_activity_count():
    specific_activities = [
        "EL-First Fix", "Installation of doors", "Waterproofing Works",
        "C-Gypsum and POP Punning", "Wall Tiling", "Floor Tiling",
        "EL-Second Fix", "Concreting", "Sewer Line", "Storm Line",
        "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone", "Electrical"
    ]
    all_activities = specific_activities + ["UP-First Fix and CP-First Fix"]

    category_mapping = {
        "EL-First Fix": "MEP",
        "UP-First Fix and CP-First Fix": "MEP",
        "C-Gypsum and POP Punning": "MEP",
        "EL-Second Fix": "MEP",
        "Concreting": "MEP",
        "Electrical": "MEP",
        "Installation of doors": "Interior Finishing",
        "Waterproofing Works": "Interior Finishing",
        "Wall Tiling": "Interior Finishing",
        "Floor Tiling": "Interior Finishing",
        "Sewer Line": "ED Civil",
        "Storm Line": "ED Civil",
        "GSB": "ED Civil",
        "WMM": "ED Civil",
        "Stamp Concrete": "ED Civil",
        "Saucer drain": "ED Civil",
        "Kerb Stone": "ED Civil"
    }

    # Enhanced activity name mapping for COS data
    cos_to_standard_mapping = {
        "No. of Slab cast": "Concreting",
        "Slab Casting": "Concreting",
        "Concrete Slab": "Concreting",
        "Slab Work": "Concreting",
        "RCC Slab": "Concreting",
        "Concreting Work": "Concreting",
        # Add more mappings as needed based on your COS data
        "Door Installation": "Installation of doors",
        "Door Fixing": "Installation of doors",
        "Waterproofing": "Waterproofing Works",
        "Wall Tiles": "Wall Tiling",
        "Floor Tiles": "Floor Tiling",
        "Electrical First Fix": "EL-First Fix",
        "Electrical Second Fix": "EL-Second Fix",
    }

    # Get Asite counts for validation
    structure_analysis = st.session_state.get('structure_analysis')
    if structure_analysis is None or structure_analysis.empty:
        st.error("❌ No structure analysis data available. Please run 'AnalyzeStatusManually' first.")
        logger.error("No structure_analysis in st.session_state for display_activity_count")
        return

    # Extract Asite counts for validation
    asite_counts = {}
    unique_towers = structure_analysis['tower_name'].unique()
    for tower in unique_towers:
        tower_data = structure_analysis[structure_analysis['tower_name'] == tower]
        tower_counts = {}
        for activity in specific_activities:
            activity_data = tower_data[tower_data['activityName'] == activity]
            count = int(activity_data['CompletedCount'].iloc[0]) if not activity_data.empty else 0
            tower_counts[activity] = count
        asite_counts[tower] = tower_counts

    count_tables = {}
    if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
        st.session_state.ai_response = {}
        logger.info("Initialized st.session_state.ai_response as empty dictionary")

    def debug_cos_data(cos_df):
        """Debug function to understand COS data structure"""
        logger.info("=== COS Data Debug Information ===")
        logger.info(f"COS DataFrame shape: {cos_df.shape}")
        logger.info(f"COS DataFrame columns: {cos_df.columns.tolist()}")
        
        # Check for activity name columns
        activity_columns = [col for col in cos_df.columns if 'activity' in col.lower() or 'name' in col.lower()]
        logger.info(f"Potential activity columns: {activity_columns}")
        
        # Check for tower/location columns
        location_columns = [col for col in cos_df.columns if any(keyword in col.lower() for keyword in ['tower', 'block', 'location', 'building'])]
        logger.info(f"Potential location columns: {location_columns}")
        
        # Sample data
        logger.debug(f"Sample COS data:\n{cos_df.head(10).to_string()}")
        
        # Check unique values in Activity Name column
        if 'Activity Name' in cos_df.columns:
            unique_activities = cos_df['Activity Name'].unique()
            logger.info(f"Unique activities in COS data: {unique_activities[:20]}")  # Show first 20
            
            # Check for concreting-related activities
            concreting_activities = [act for act in unique_activities if act and ('slab' in str(act).lower() or 'concrete' in str(act).lower())]
            logger.info(f"Concreting-related activities found: {concreting_activities}")

    def improved_tower_mapping(cos_df):
        """Improved tower mapping logic"""
        tower_dfs = {'EWS Tower 1': pd.DataFrame(), 'LIG Tower 3': pd.DataFrame()}
        
        # Debug the COS data first
        debug_cos_data(cos_df)
        
        # Try multiple approaches for tower mapping
        
        # Approach 1: Use qiLocationId if available
        structure_locations = st.session_state.get('structure_location_data')
        if structure_locations is not None and 'qiLocationId' in cos_df.columns:
            logger.info("Using qiLocationId approach for tower mapping")
            parent_child_dict = dict(zip(structure_locations['qiLocationId'], structure_locations['qiParentId']))
            name_dict = dict(zip(structure_locations['qiLocationId'], structure_locations['name']))
            cos_df['full_path'] = cos_df['qiLocationId'].apply(
                lambda x: get_full_path(x, parent_child_dict, name_dict)
            )
            cos_df['tower_name'] = cos_df['full_path'].apply(get_tower_name)
            
            for tower in ['EWS Tower 1', 'LIG Tower 3']:
                tower_data = cos_df[cos_df['tower_name'] == tower].copy()
                tower_dfs[tower] = tower_data
                logger.info(f"qiLocationId approach - Data for {tower}: {len(tower_data)} rows")
        
        # Approach 2: Check for direct tower columns
        else:
            tower_columns = ['Tower', 'Block', 'tower_name', 'Building', 'Structure']
            tower_col = None
            for col in tower_columns:
                if col in cos_df.columns:
                    tower_col = col
                    logger.info(f"Found tower column: {col}")
                    break
            
            if tower_col:
                logger.info(f"Using {tower_col} column for tower mapping")
                # Check unique values in tower column
                unique_towers_in_data = cos_df[tower_col].unique()
                logger.info(f"Unique values in {tower_col}: {unique_towers_in_data}")
                
                for tower in ['EWS Tower 1', 'LIG Tower 3']:
                    # Try exact match first
                    tower_data = cos_df[cos_df[tower_col] == tower].copy()
                    if tower_data.empty:
                        # Try partial matches
                        if 'EWS' in tower:
                            tower_data = cos_df[cos_df[tower_col].str.contains('EWS', case=False, na=False)].copy()
                        elif 'LIG' in tower:
                            tower_data = cos_df[cos_df[tower_col].str.contains('LIG', case=False, na=False)].copy()
                    
                    tower_dfs[tower] = tower_data
                    logger.info(f"Tower column approach - Data for {tower}: {len(tower_data)} rows")
            
            # Approach 3: Infer from Activity Name or other text fields
            else:
                logger.info("Using Activity Name inference for tower mapping")
                text_columns = ['Activity Name', 'Description', 'Location', 'Notes']
                available_text_cols = [col for col in text_columns if col in cos_df.columns]
                
                def infer_tower_from_text(row):
                    text_to_check = ""
                    for col in available_text_cols:
                        if pd.notna(row[col]):
                            text_to_check += str(row[col]).lower() + " "
                    
                    if 'ews' in text_to_check or 'tower 1' in text_to_check:
                        return 'EWS Tower 1'
                    elif 'lig' in text_to_check or 'tower 3' in text_to_check:
                        return 'LIG Tower 3'
                    return None
                
                cos_df['inferred_tower'] = cos_df.apply(infer_tower_from_text, axis=1)
                
                for tower in ['EWS Tower 1', 'LIG Tower 3']:
                    tower_data = cos_df[cos_df['inferred_tower'] == tower].copy()
                    tower_dfs[tower] = tower_data
                    logger.info(f"Text inference approach - Data for {tower}: {len(tower_data)} rows")
        
        return tower_dfs

    def process_tower_data(tower_data, tname):
        if tower_data is None or tower_data.empty:
            logger.warning(f"No COS data available for {tname}. Using Asite data as fallback.")
            # Create count table with Asite data
            count_table = pd.DataFrame({
                'Count_Unfiltered': [asite_counts.get(tname, {}).get(activity, 0) for activity in all_activities],
                'Count_Filtered': [asite_counts.get(tname, {}).get(activity, 0) for activity in all_activities]
            }, index=all_activities)
            return tname, count_table

        tower_data = tower_data.copy()
        logger.debug(f"Processing {tname} with {len(tower_data)} rows")
        logger.debug(f"Processing {tname} with columns: {tower_data.columns.tolist()}")

        # Enhanced activity name mapping
        if 'Activity Name' in tower_data.columns:
            # Log original activity names
            original_activities = tower_data['Activity Name'].unique()
            logger.info(f"Original activities for {tname}: {original_activities}")
            
            # Apply mapping
            tower_data['Activity Name'] = tower_data['Activity Name'].map(
                lambda x: cos_to_standard_mapping.get(x, x) if pd.notna(x) else x
            )
            
            # Log mapped activity names
            mapped_activities = tower_data['Activity Name'].unique()
            logger.info(f"Mapped activities for {tname}: {mapped_activities}")

        count_table = pd.DataFrame({
            'Count_Unfiltered': [0] * len(all_activities),
            'Count_Filtered': [0] * len(all_activities)
        }, index=all_activities)

        # Process Actual Finish filtering
        if 'Actual Finish' in tower_data.columns:
            tower_data['Actual_Finish_Original'] = tower_data['Actual Finish'].astype(str)
            tower_data['Actual Finish'] = pd.to_datetime(tower_data['Actual Finish'], errors='coerce')
            has_na_mask = (
                pd.isna(tower_data['Actual Finish']) |
                (tower_data['Actual_Finish_Original'].str.upper() == 'NAT') |
                (tower_data['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
            )
            tower_data_filtered = tower_data[~has_na_mask].copy()
            tower_data.drop('Actual_Finish_Original', axis=1, inplace=True)
        else:
            logger.warning(f"No 'Actual Finish' column in {tname}. Using unfiltered data.")
            tower_data_filtered = tower_data.copy()

        # Count activities
        for activity in specific_activities:
            if 'Activity Name' in tower_data.columns:
                exact_matches = tower_data[tower_data['Activity Name'] == activity]
                count_table.loc[activity, 'Count_Unfiltered'] = len(exact_matches)
                exact_matches_filtered = tower_data_filtered[tower_data_filtered['Activity Name'] == activity]
                count_table.loc[activity, 'Count_Filtered'] = len(exact_matches_filtered)
                
                # Log specific counts for debugging
                if activity == 'Concreting' and len(exact_matches) > 0:
                    logger.info(f"Found {len(exact_matches)} Concreting activities for {tname}")
                    logger.debug(f"Concreting activities: {exact_matches['Activity Name'].tolist()}")

        # Handle UP-First Fix and CP-First Fix combination
        if 'Activity Name' in tower_data.columns:
            up_first_fix_matches = tower_data[tower_data['Activity Name'] == "UP-First Fix"]
            cp_first_fix_matches = tower_data[tower_data['Activity Name'] == "CP-First Fix"]
            up_first_fix_count = len(up_first_fix_matches)
            cp_first_fix_count = len(cp_first_fix_matches)
            count_table.loc["UP-First Fix and CP-First Fix", "Count_Unfiltered"] = min(up_first_fix_count, cp_first_fix_count)

            up_first_fix_matches_filtered = tower_data_filtered[tower_data_filtered['Activity Name'] == "UP-First Fix"]
            cp_first_fix_matches_filtered = tower_data_filtered[tower_data_filtered['Activity Name'] == "CP-First Fix"]
            up_first_fix_count_filtered = len(up_first_fix_matches_filtered)
            cp_first_fix_count_filtered = len(cp_first_fix_matches_filtered)
            count_table.loc["UP-First Fix and CP-First Fix", "Count_Filtered"] = min(up_first_fix_count_filtered, cp_first_fix_count_filtered)

        # If COS counts are zero but Asite has data, use Asite as fallback
        for activity in specific_activities:
            cos_count = count_table.loc[activity, 'Count_Filtered']
            asite_count = asite_counts.get(tname, {}).get(activity, 0)
            
            if cos_count == 0 and asite_count > 0:
                logger.warning(f"COS count for {activity} in {tname} is 0, but Asite count is {asite_count}. Using Asite count.")
                count_table.loc[activity, 'Count_Filtered'] = asite_count
                count_table.loc[activity, 'Count_Unfiltered'] = asite_count

        count_table['Count_Unfiltered'] = count_table['Count_Unfiltered'].astype(int)
        count_table['Count_Filtered'] = count_table['Count_Filtered'].astype(int)

        logger.info(f"Final counts for {tname}:")
        logger.info(f"Concreting: Unfiltered = {count_table.loc['Concreting', 'Count_Unfiltered']}, Filtered = {count_table.loc['Concreting', 'Count_Filtered']}")

        return tname, count_table

    # Main processing logic
    if not unique_towers.size:
        st.error("❌ No towers found in structure analysis data.")
        logger.error("No unique towers in structure_analysis")
        return

    # Get COS data
    cos_df = st.session_state.get('cos_df_Revised_Baseline_45daysNGT_Rai')
    if cos_df is None or cos_df.empty:
        st.warning("⚠️ No COS data available. Using Asite data only.")
        logger.warning("No COS data available, using Asite data as fallback")
        
        # Use Asite data only
        for tower in ['EWS Tower 1', 'LIG Tower 3']:
            if tower in asite_counts:
                count_table = pd.DataFrame({
                    'Count_Unfiltered': [asite_counts[tower].get(activity, 0) for activity in all_activities],
                    'Count_Filtered': [asite_counts[tower].get(activity, 0) for activity in all_activities]
                }, index=all_activities)
                count_tables[tower] = count_table
    else:
        # Process COS data with improved mapping
        tower_dfs = improved_tower_mapping(cos_df)
        
        # Process each tower
        for tower in ['EWS Tower 1', 'LIG Tower 3']:
            tower_data = tower_dfs.get(tower)
            tname, count_table = process_tower_data(tower_data, tower)
            if count_table is not None:
                count_tables[tower] = count_table

    if not count_tables:
        st.error("No data available for EWS Tower 1 or LIG Tower 3.")
        logger.error("No count tables generated for towers")
        return

    # Display results
    for tname, count_table in count_tables.items():
        with st.spinner(f"Processing activity counts for {tname}..."):
            try:
                st.write(f"Activity Count for {tname}:")
                st.write(count_table)

                count_table_filtered = count_table[['Count_Filtered']].rename(columns={'Count_Filtered': 'Count'})
                logger.debug(f"Data sent to WatsonX for {tname}:\n{count_table_filtered.to_string()}")
                
                ai_response = generatePrompt(count_table_filtered, tname)
                ai_data = json.loads(ai_response)

                st.session_state.ai_response[tname] = ai_data
                logger.info(f"Stored AI response for {tname}")

                totals_mapping = getTotal(ai_data)

                # Validate against Asite data and show comparison
                for activity in specific_activities:
                    cos_count = totals_mapping.get(activity, count_table.loc[activity, 'Count_Filtered'])
                    asite_count = asite_counts.get(tname, {}).get(activity, 0)
                    
                    if cos_count != asite_count and asite_count > 0:
                        if cos_count == 0:
                            st.warning(f"⚠️ {activity} count mismatch for {tname}: COS count = {cos_count}, Asite count = {asite_count}. Using Asite count.")
                            totals_mapping[activity] = asite_count
                        else:
                            st.info(f"ℹ️ {activity} count difference for {tname}: COS count = {cos_count}, Asite count = {asite_count}. Using COS count.")

                # Create display DataFrame
                display_df = count_table.reset_index().rename(columns={'index': 'Activity Name'})
                display_df['Total'] = display_df['Activity Name'].map(
                    lambda x: totals_mapping.get(x, display_df.loc[display_df['Activity Name'] == x, 'Count_Filtered'].iloc[0])
                )
                display_df['Category'] = display_df['Activity Name'].map(category_mapping)
                display_df['Asite_Count'] = display_df['Activity Name'].map(
                    lambda x: asite_counts.get(tname, {}).get(x, 0)
                )

                st.write(f"Activity Count with Totals for {tname}:")
                st.write(display_df[['Activity Name', 'Count_Filtered', 'Total', 'Asite_Count', 'Category']])

                # Display by category
                for category in ['MEP', 'Interior Finishing', 'ED Civil', 'Structure Work']:
                    category_df = display_df[display_df['Category'] == category]
                    if not category_df.empty:
                        st.write(f"**{category} ({tname})**")
                        st.write(category_df[['Activity Name', 'Count_Filtered', 'Total', 'Asite_Count']])

            except Exception as e:
                logger.error(f"Error processing {tname}: {str(e)}")
                st.error(f"Error processing {tname}: {str(e)}")
                # Generate fallback using Asite data
                st.info(f"Using Asite data as fallback for {tname}")
                count_table_asite = pd.DataFrame({
                    'Count_Filtered': [asite_counts.get(tname, {}).get(activity, 0) for activity in all_activities]
                }, index=all_activities)
                ai_response = generate_fallback_totals(count_table_asite)



# Combined function for Initialize and Fetch Data
@function_timer()
async def initialize_and_fetch_data(email, password):
    with st.spinner("Starting initialization and data fetching process..."):
        # Step 1: Login
        if not email or not password:
            st.sidebar.error("Please provide both email and password!")
            return False
        try:
            st.sidebar.write("Logging in...")
            session_id = await login_to_asite(email, password)
            if not session_id:
                st.sidebar.error("Login failed!")
                return False
            st.sidebar.success("Login successful!")
        except Exception as e:
            st.sidebar.error(f"Login failed: {str(e)}")
            return False

        # Step 2: Get Workspace ID
        try:
            st.sidebar.write("Fetching Workspace ID...")
            await GetWorkspaceID()
            st.sidebar.success("Workspace ID fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
            return False

        # Step 3: Get Project IDs
        try:
            st.sidebar.write("Fetching Project IDs...")
            await GetProjectId()
            st.sidebar.success("Project IDs fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
            return False

        # Step 4: Get All Data (Structure only)
        try:
            st.sidebar.write("Fetching All Data...")
            Edenstructure = await GetAllDatas()
            st.session_state.eden_structure = Edenstructure
            st.sidebar.success("All Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
            return False

        # Step 5: Get Activity Data
        try:
            st.sidebar.write("Fetching Activity Data...")
            structure_activity_data = await Get_Activity()
            st.session_state.structure_activity_data = structure_activity_data
            st.sidebar.success("Activity Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
            return False

        # Step 6: Get Location/Module Data
        try:
            st.sidebar.write("Fetching Location/Module Data...")
            structure_location_data = await Get_Location()
            st.session_state.structure_location_data = structure_location_data 
            st.sidebar.success("Location/Module Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
            return False

        # Step 7: Fetch COS Files
        try:
            st.sidebar.write("Fetching COS files from Wave City Club folder...")
            file_key = get_cos_files()
            st.session_state.file_key = file_key
            if file_key:
                st.success(f"Found 1 file in COS storage: {file_key}")
                try:
                    st.write(f"Processing file: {file_key}")
                    cos_client = initialize_cos_client()
                    if not cos_client:
                        st.error("Failed to initialize COS client during file fetch")
                        logger.error("COS client initialization failed during file fetch")
                        return False
                    st.write("Fetching file from COS...")
                    response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
                    file_bytes = io.BytesIO(response['Body'].read())
                    st.write("File fetched successfully. Processing sheets...")
                    results = process_file(file_bytes, file_key)
                    st.write(f"Processing results: {len(results)} sheets processed")
                    for df, sheet_name in results:
                        if df is not None:
                            if sheet_name == "Revised Baseline 45daysNGT+Rai":  
                                st.session_state.cos_df_Revised_Baseline_45daysNGT_Rai = df
                                st.session_state.cos_tname_Revised_Baseline_45daysNGT_Rai = "Revised Baseline 45daysNGT+Rai"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                        else:
                            st.warning(f"No data processed for {sheet_name} in {file_key}.")
                except Exception as e:
                    st.error(f"Error loading {file_key} from cloud storage: {str(e)}")
                    logger.error(f"Error loading {file_key}: {str(e)}")
                    return False
            else:
                st.warning("No expected Excel files available in the 'Wave City Club' folder of the COS bucket.")
                return False
        except Exception as e:
            st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
            logger.error(f"Failed to fetch COS files: {str(e)}")
            return False

    st.sidebar.success("All steps completed successfully!")
    return True


@function_timer()
def generate_consolidated_Checklist_excel(structure_analysis=None, activity_counts=None):
    try:
        # Add validation at the beginning
        if structure_analysis is None:
            structure_analysis = st.session_state.get('structure_analysis', None)
            if structure_analysis is None:
                st.error("❌ No structure analysis data available.")
                logger.error("structure_analysis is None in generate_consolidated_Checklist_excel")
                return None
        
        if activity_counts is None:
            activity_counts = st.session_state.get('ai_response', {})
            if not activity_counts:
                st.error("❌ No activity counts data available.")
                logger.error("activity_counts is empty in generate_consolidated_Checklist_excel")
                return None

        # Validate structure_analysis columns
        if not isinstance(structure_analysis, pd.DataFrame):
            st.error("❌ structure_analysis is not a DataFrame.")
            logger.error("structure_analysis is not a DataFrame")
            return None

        expected_columns = ['tower_name', 'activityName', 'CompletedCount']
        missing_columns = [col for col in expected_columns if col not in structure_analysis.columns]
        if missing_columns:
            st.error(f"❌ Missing columns in structure_analysis: {missing_columns}")
            logger.error(f"Missing columns in structure_analysis: {missing_columns}")
            return None

        # Transform activity_counts if it's a dictionary
        transformed_activity_counts = []
        if isinstance(activity_counts, dict):
            for tower, categories_data in activity_counts.items():
                for category_data in categories_data:
                    for activity_data in category_data.get("Activities", []):
                        transformed_activity_counts.append({
                            "tower": tower,
                            "activity": activity_data.get("Activity Name"),
                            "completed_count": activity_data.get("Total", 0)
                        })
        else:
            transformed_activity_counts = activity_counts

        # Validate transformed_activity_counts
        if not isinstance(transformed_activity_counts, list):
            st.error("❌ Transformed activity_counts is not a list.")
            logger.error("Transformed activity_counts is not a list")
            return None

        # Define categories and mappings
        categories = {
            "Interior Finishing (Civil)": ["Installation of doors", "Waterproofing Works", "Wall Tiling", "Floor Tiling"],
            "MEP": ["EL-First Fix", "Plumbing Works", "C-Gypsum and POP P Suddenning", "EL-Second Fix", "Concreting", "Electrical"],
            "External Development (Civil)": ["Sewer Line", "Storm Line", "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone"]
        }

        cos_to_asite_mapping = {
            "EL-First Fix": "Wall Conducting",
            "Installation of doors": ["Door/Window Frame", "Door/Window Shutter"],
            "Plumbing Works": "Plumbing Works",
            "Waterproofing Works": "Waterproofing - Sunken",
            "C-Gypsum and POP P Suddenning": "POP & Gypsum Plaster",
            "Wall Tiling": "Wall Tile",
            "Floor Tiling": "Floor Tiling",
            "EL-Second Fix": "Wiring & Switch Socket",
            "Concreting": "Concreting",
            "Sewer Line": "Sewer Line",
            "Storm Line": "Rain Water/Storm",
            "GSB": "Granular Sub-base",
            "WMM": "WMM",
            "Saucer drain": "Saucer drain/Paver block",
            "Kerb Stone": "Kerb Stone",  # Ensured mapping is consistent
            "Electrical": "Electrical Cable"
        }

        # Define towers
        towers = ["EWS Tower 1", "LIG Tower 3"]

        # Ensure slabreport is populated
        if "slabreport" not in st.session_state or not st.session_state.slabreport:
            st.write("Fetching slab report data...")
            GetSlabReport()  # This will populate st.session_state.slabreport

        # Parse slabreport to extract tracker counts
        try:
            if isinstance(st.session_state.slabreport, str) and st.session_state.slabreport == "No Data Found":
                st.error("❌ No slab report data found in COS storage.")
                logger.error("No slab report data found in st.session_state.slabreport")
                return None

            slab_data = json.loads(st.session_state.slabreport) if isinstance(st.session_state.slabreport, str) else st.session_state.slabreport
            if not isinstance(slab_data, list):
                st.error("❌ Invalid slab report data format: expected a list.")
                logger.error(f"Invalid slab report data format: {type(slab_data)}")
                return None
        except json.JSONDecodeError as e:
            st.error("❌ Failed to parse slab report data: invalid JSON.")
            logger.error(f"Failed to parse slab report data: {str(e)}")
            return None
        except Exception as e:
            st.error("❌ Error processing slab report data.")
            logger.error(f"Error processing slab report data: {str(e)}")
            return None

        # Map slab report data to tracker_counts format
        # Assuming slab_data is in the format [{"Tower": "EWST1", "Slab Count": 51}, ...]
        tracker_counts = []
        expected_towers = ["EWST1", "EWST2", "EWST3", "LIGT1", "LIGT2", "LIGT3"]
        tower_counts = {tower: 0 for tower in expected_towers}  # Default to 0 for all towers

        for entry in slab_data:
            tower = entry.get("Tower")
            # The key for the count might be "Slab Count", "Green (1)", or something else depending on ProcessEWS_LIG
            # We'll check common possibilities
            count = entry.get("Slab Count") or entry.get("Green (1)") or entry.get("Count") or 0
            if tower in expected_towers:
                tower_counts[tower] = int(count) if isinstance(count, (int, float)) and not pd.isna(count) else 0

        # Convert to tracker_counts format
        for tower, count in tower_counts.items():
            tracker_counts.append({"Tower": tower, "Green (1)": count})

        # Map tracker tower names to function tower names
        tower_mapping = {
            "EWST1": "EWS Tower 1",
            "EWST2": "EWS Tower 2",
            "EWST3": "EWS Tower 3",
            "LIGT1": "LIG Tower 1",
            "LIGT2": "LIG Tower 2",
            "LIGT3": "LIG Tower 3"
        }

        # Create a dictionary of closed checklist counts from tracker_counts
        closed_counts = {}
        for entry in tracker_counts:
            tracker_tower = entry.get("Tower")
            green_count = entry.get("Green (1)", 0)
            mapped_tower = tower_mapping.get(tracker_tower)
            if mapped_tower:
                closed_counts[mapped_tower] = green_count

        consolidated_rows = []

        for tower in towers:
            tower_key = tower
            # Get the closed checklist count for this tower from tracker_counts
            tower_closed_count = closed_counts.get(tower, 0)

            for category, activities in categories.items():
                if not activities and "Structure" not in category:
                    continue

                if activities:
                    for activity in activities:
                        asite_activity = cos_to_asite_mapping.get(activity, activity)
                        if isinstance(asite_activity, list):
                            asite_activities = asite_activity
                        else:
                            asite_activities = [asite_activity]

                        # Set closed checklist to tower_closed_count only for "Concreting"
                        if activity == "Concreting":
                            closed_checklist = tower_closed_count
                        else:
                            closed_checklist = 0  # Default to 0 for all other activities

                        completed_flats = 0
                        # Use transformed_activity_counts for completed flats
                        tower_data = [item for item in transformed_activity_counts if item.get("tower") == tower]
                        if tower_data:
                            if activity == "Plumbing Works":
                                for item in tower_data:
                                    if item.get("activity") == "Min. count of UP-First Fix and CP-First Fix":
                                        completed_flats = item.get("completed_count", 0)
                                        break
                            else:
                                for item in tower_data:
                                    if item.get("activity") == activity:
                                        completed_flats = item.get("completed_count", 0)
                                        break

                        in_progress = 0
                        # Calculate Open/Missing check list, ensuring positive value
                        open_missing = abs(completed_flats - closed_checklist)

                        display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

                        consolidated_rows.append({
                            "Tower": tower,
                            "Category": category,
                            "Activity Name": display_activity,
                            "Completed Work*(Count of Flat)": completed_flats,
                            "In progress": in_progress,
                            "Closed checklist": closed_checklist,
                            "Open/Missing check list": open_missing
                        })
                else:
                    # For categories with no activities (e.g., "Structure"), set closed_checklist to 0
                    consolidated_rows.append({
                        "Tower": tower,
                        "Category": category,
                        "Activity Name": "",
                        "Completed Work*(Count of Flat)": 0,
                        "In progress": 0,
                        "Closed checklist": 0,  # Set to 0 since this isn't "Concreting"
                        "Open/Missing check list": 0
                    })

        df = pd.DataFrame(consolidated_rows)
        if df.empty:
            st.warning("No data available to generate consolidated checklist.")
            return None

        df.sort_values(by=["Tower", "Category"], inplace=True)

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Consolidated Checklist")

        header_format = workbook.add_format({
            'bold': True, 
            'bg_color': '#D3D3D3',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        total_format = workbook.add_format({
            'bold': True, 
            'bg_color': '#FFDAB9',
            'border': 1,
            'align': 'center'
        })
        cell_format = workbook.add_format({'border': 1})

        headers = ["Activity Name", "Completed", "In progress", "Closed checklist", "Open/Missing check list"]

        grouped_by_tower = df.groupby('Tower')
        
        current_row = 0
        
        for tower, tower_group in grouped_by_tower:
            grouped_by_category = tower_group.groupby('Category')
            categories = list(grouped_by_category.groups.keys())
            categories_per_row = 2
            col_width = 6
            
            for i, (category, cat_group) in enumerate(grouped_by_category):
                row_offset = (i // categories_per_row) * 12
                col_offset = (i % categories_per_row) * col_width
                table_start_row = current_row + row_offset
                table_start_col = col_offset
                
                category_title = f"{tower} {category} Checklist Status"
                worksheet.merge_range(
                    table_start_row, table_start_col, 
                    table_start_row, table_start_col + 4, 
                    category_title, header_format
                )
                
                for j, header in enumerate(headers):
                    worksheet.write(table_start_row + 1, table_start_col + j, header, header_format)
                
                data_start_row = table_start_row + 2
                total_pending = 0
                
                if not cat_group.empty and cat_group["Activity Name"].iloc[0] != "":
                    for idx, (_, row) in enumerate(cat_group.iterrows()):
                        current_data_row = data_start_row + idx
                        worksheet.write(current_data_row, table_start_col + 0, row["Activity Name"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 1, row["Completed Work*(Count of Flat)"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 2, row["In progress"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 3, row["Closed checklist"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 4, row["Open/Missing check list"], cell_format)
                        total_pending += row["Open/Missing check list"]  # Sum the positive open/missing values
                    
                    min_rows = 5
                    actual_rows = len(cat_group)
                    for empty_row in range(actual_rows, min_rows):
                        current_data_row = data_start_row + empty_row
                        for col in range(5):
                            worksheet.write(current_data_row, table_start_col + col, "", cell_format)
                else:
                    for empty_row in range(5):
                        current_data_row = data_start_row + empty_row
                        for col in range(5):
                            worksheet.write(current_data_row, table_start_col + col, "", cell_format)
                
                total_row = data_start_row + 5
                worksheet.merge_range(
                    total_row, table_start_col, 
                    total_row, table_start_col + 3, 
                    "Total pending check list", total_format
                )
                worksheet.write(total_row, table_start_col + 4, total_pending, total_format)
            
            max_categories_in_row = (len(categories) + categories_per_row - 1) // categories_per_row
            current_row += max_categories_in_row * 12 + 3
        
        for col in range(12):
            worksheet.set_column(col, col, 18)

        workbook.close()
        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating consolidated Excel: {str(e)}")
        st.error(f"❌ Error generating Excel file: {str(e)}")
        return None

# Combined function to handle analysis and display
@function_timer(show_args=True)
def run_analysis_and_display():
    try:
        # Step 1: Run status analysis
        st.write("Running status analysis...")
      
        AnalyzeStatusManually()
        
        # Check if structure_analysis was populated
        if 'structure_analysis' not in st.session_state or st.session_state.structure_analysis is None:
            st.error("❌ Status analysis failed to generate structure_analysis. Please check the logs and ensure data fetching was successful.")
            logger.error("run_analysis_and_display failed: structure_analysis not populated after AnalyzeStatusManually")
            return
        st.success("Status analysis completed successfully!")

        # Step 2: Initialize AI response if needed
        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
            logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

        # Step 3: Display activity counts
        st.write("Displaying activity counts and generating AI data...")
        logger.debug("COS DataFrame columns: {}".format(
            list(st.session_state.get('cos_df_Revised_Baseline_45daysNGT_Rai', pd.DataFrame()).columns)
        ))
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        
        
        # Check structure_analysis
        structure_analysis = st.session_state.get('structure_analysis')
        if structure_analysis is None:
            st.error("❌ No structure analysis data available.")
            logger.error("No structure_analysis in st.session_state")
            return
        
        # Check ai_response (activity_counts)
        if not st.session_state.ai_response:
            st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
            logger.error("No AI data in st.session_state.ai_response after display_activity_count")
            
            logger.debug("Retrying COS DataFrame columns: {}".format(
                list(st.session_state.get('cos_df_Revised_Baseline_45daysNGT_Rai', pd.DataFrame()).columns)
            ))
            display_activity_count()
            if not st.session_state.ai_response:
                st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
                logger.error("Failed to regenerate AI data")
                return

     
        # Step 6: Generate Excel file
        st.write("Generating consolidated checklist Excel file...")
        
        with st.spinner("Generating Excel file... This may take a moment."):
            excel_file = generate_consolidated_Checklist_excel(
                structure_analysis=structure_analysis, 
                activity_counts=st.session_state.ai_response
            )
        
        # Step 7: Handle download
        if excel_file:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_EWS_LIG_{timestamp}.xlsx"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.download_button(
                    label="📥 Download Checklist Excel",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button",
                    help="Click to download the consolidated checklist in Excel format."
                )
            st.success("✅ Excel file generated successfully!")
        else:
            st.error("Error generating Excel file. Check logs for details.")
            logger.error("Failed to generate Excel file - function returned None")

    except Exception as e:
        error_msg = str(e)
        st.error(f"❌ Error during analysis, display, or Excel generation: {error_msg}")
        logger.error(f"Error during analysis, display, or Excel generation: {error_msg}")
        logger.error(f"Stack trace:\n{traceback.format_exc()}")
            
# Streamlit UI
st.markdown(
    """
    <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
               color: red; 
               font-size: 48px; 
               text-align: center;'>
        CheckList - Report
    </h1>
    """,
    unsafe_allow_html=True
)

# Initialize and Fetch Data
st.sidebar.title("🔒 Asite Initialization")
email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

if st.sidebar.button("Initialize and Fetch Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        success = loop.run_until_complete(initialize_and_fetch_data(email, password))
        if success:
            st.sidebar.success("Initialization and data fetching completed successfully!")
        else:
            st.sidebar.error("Initialization and data fetching failed!")
    except Exception as e:
        st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
    finally:
        loop.close()

# Analyze and Display
st.sidebar.title("📊 Status Analysis")
if st.sidebar.button("Analyze and Display Activity Counts"):
    try:
        run_analysis_and_display()  # This function already handles the full workflow
    except Exception as e:
        logging.error(f"Error during analysis and display: {str(e)}")
        logging.error(f"Stack trace:\n{traceback.format_exc()}")
        st.error(f"Error occurred: {str(e)}\nCheck logs for details.")

st.sidebar.title("📊 Slab Cycle")
st.session_state.ignore_year = st.sidebar.number_input("Ignore Year", min_value=1900, max_value=2100, value=2023, step=1, key="ignore_year1")
st.session_state.ignore_month = st.sidebar.number_input("Ignore Month", min_value=1, max_value=12, value=3, step=1, key="ignore_month1")


































