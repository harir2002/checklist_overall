
import streamlit as st
import requests
import json
import urllib.parse
import urllib3
import certifi
import pandas as pd
from datetime import datetime
import re
import logging
import os
from dotenv import load_dotenv
import aiohttp
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import openpyxl
import io
from dotenv import load_dotenv
from uuid import uuid4
import ibm_boto3
from ibm_botocore.client import Config
from tenacity import retry, stop_after_attempt, wait_exponential
import xlsxwriter

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# IBM COS Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
COS_BUCKET = os.getenv("COS_BUCKET")

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")

# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

# Login Function
async def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
    if response.status_code == 200:
        try:
            session_id = response.json().get("UserProfile", {}).get("Sessionid")
            logger.info(f"Login successful, Session ID: {session_id}")
            st.session_state.sessionid = session_id
            st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
            return session_id
        except json.JSONDecodeError:
            logger.error("JSONDecodeError during login")
            st.sidebar.error("❌ Failed to parse login response")
            return None
    logger.error(f"Login failed: {response.status_code} - {response.text}")
    st.sidebar.error(f"❌ Login failed: {response.status_code} - {response.text}")
    return None

# Function to generate access token
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
def get_access_token(API_KEY):
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
    response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
    try:
        if response.status_code == 200:
            token_info = response.json()
            logger.info("Access token generated successfully")
            return token_info['access_token']
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
            raise Exception("Failed to get access token")
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f"❌ Error getting access token: {str(e)}")
        return None

# Initialize COS client
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
def initialize_cos_client():
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        logger.info("COS client initialized successfully")
        return cos_client
    except Exception as e:
        logger.error(f"Error initializing COS client: {str(e)}")
        st.error(f"❌ Error initializing COS client: {str(e)}")
        raise

# Fetch Workspace ID
async def GetWorkspaceID():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch workspace list: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch workspace list: {response.status_code}")
    try:
        data = response.json()
        st.session_state.workspaceid = data['asiteDataList']['workspaceVO'][3]['Workspace_Id']
        st.write(f"Workspace ID: {st.session_state.workspaceid}")
    except (KeyError, IndexError) as e:
        st.error(f"Error parsing workspace ID: {str(e)}")
        raise

# Fetch Project IDs
async def GetProjectId():
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch project IDs: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch project IDs: {response.status_code}")
    data = response.json()
    if not data.get('data'):
        st.error("No quality plans found for the specified date.")
        raise Exception("No quality plans found")
    st.session_state.Wave_City_Club_structure = data['data'][0]['planId']
    st.write(f"Wave City Club Structure Project ID: {st.session_state.Wave_City_Club_structure}")

# Asynchronous Fetch Function
async def fetch_data(session, url, headers):
    async with session.get(url, headers=headers) as response:
        if response.status == 200:
            return await response.json()
        elif response.status == 204:
            return None
        else:
            raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# Fetch All Structure Data
async def GetAllDatas():
    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_structure_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Wave_City_Club Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                async with session.get(url, headers=headers) as response:
                    if response.status == 204:
                        st.write("No more Wave_City_Club Structure data available (204)")
                        break
                    data = await response.json()
                    if 'associationList' in data and data['associationList']:
                        all_structure_data.extend(data['associationList'])
                    else:
                        all_structure_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_structure_data[-record_limit:])} Wave_City_Club Structure records (Total: {len(all_structure_data)})")
                    if len(all_structure_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure data: {str(e)}")
                break

    df_structure = pd.DataFrame(all_structure_data)
    
    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_structure.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_structure.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("❌ Neither statusName nor statusColor found in data!")
        return pd.DataFrame()

    Wave_City_Club_structure = df_structure[desired_columns]

    st.write(f"Wave_City_Club STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(Wave_City_Club_structure)}")
    st.write(Wave_City_Club_structure)
    
    return Wave_City_Club_structure

# Fetch Activity Data
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_activity_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Activity data for Wave_City_Club Structure...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
                break
 
    structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

    st.write("Wave_City_Club STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)
      
    return structure_activity_data

# Fetch Location/Module Data
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_location_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Wave_City_Club Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Location data: {str(e)}")
                break
        
    structure_df = pd.DataFrame(all_structure_location_data)
    
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("❌ All 'name' values in Structure Location data are missing or empty!")

    st.write("Wave_City_Club STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)
    
    st.session_state.structure_location_data = structure_df
    
    return structure_df

# Process individual chunk
def process_chunk(chunk, chunk_idx, dataset_name, location_df):
    logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
    generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
    logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
    return generated_text, chunk_idx

# Process data with manual counting
def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return "No completed activities found."

    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved Wave_City_Club {dataset_name} data to Wave_City_Club_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    location_df = st.session_state.structure_location_data

    chunk_results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_chunk = {
            executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx 
            for idx, chunk in enumerate(chunks)
        }

        completed_chunks = 0
        for future in as_completed(future_to_chunk):
            chunk_idx = future_to_chunk[future]
            try:
                generated_text, idx = future.result()
                chunk_results[idx] = generated_text
                completed_chunks += 1
                progress_percent = completed_chunks / len(chunks)
                progress_bar.progress(progress_percent)
                status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
                st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

    parsed_data = {}
    for chunk_idx in sorted(chunk_results.keys()):
        generated_text = chunk_results[chunk_idx]
        if not generated_text:
            continue

        current_tower = None
        tower_activities = []
        lines = generated_text.split("\n")
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            if line.startswith("Tower:"):
                try:
                    tower_parts = line.split("Tower:", 1)
                    if len(tower_parts) > 1:
                        if current_tower and tower_activities:
                            if current_tower not in parsed_data:
                                parsed_data[current_tower] = []
                            parsed_data[current_tower].extend(tower_activities)
                            tower_activities = []
                        current_tower = tower_parts[1].strip()
                except Exception as e:
                    logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
                    if not current_tower:
                        current_tower = f"Unknown Tower {chunk_idx}"
                    
            elif line.startswith("Total Completed Activities:"):
                continue
            elif not line.strip().startswith("activityName"):
                try:
                    parts = re.split(r'\s{2,}', line.strip())
                    if len(parts) >= 2:
                        activity_name = ' '.join(parts[:-1]).strip()
                        count_str = parts[-1].strip()
                        count_match = re.search(r'\d+', count_str)
                        if count_match:
                            count = int(count_match.group())
                            if current_tower:
                                tower_activities.append({
                                    "activityName": activity_name,
                                    "completedCount": count
                                })
                    else:
                        match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
                        if match and current_tower:
                            activity_name = match.group(1).strip()
                            count = int(match.group(2).strip())
                            tower_activities.append({
                                "activityName": activity_name,
                                "completedCount": count
                            })
                except (ValueError, IndexError) as e:
                    logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)

    aggregated_data = {}
    for tower_name, activities in parsed_data.items():
        tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
        if tower_short_name not in aggregated_data:
            aggregated_data[tower_short_name] = {}
        
        for activity in activities:
            name = activity.get("activityName", "Unknown")
            count = activity.get("completedCount", 0)
            if name in aggregated_data[tower_short_name]:
                aggregated_data[tower_short_name][name] += count
            else:
                aggregated_data[tower_short_name][name] = count

    combined_output_lines = []
    sorted_towers = sorted(aggregated_data.keys())
    
    for i, tower_short_name in enumerate(sorted_towers):
        combined_output_lines.append(f"{tower_short_name:<11} activityName            CompletedCount")
        activity_dict = aggregated_data[tower_short_name]
        tower_total = 0
        for name, count in sorted(activity_dict.items()):
            combined_output_lines.append(f"{'':<11} {name:<23} {count:>14}")
            tower_total += count
        combined_output_lines.append(f"{'':<11} Total for {tower_short_name:<11}: {tower_total:>14}")
        if i < len(sorted_towers) - 1:
            combined_output_lines.append("")
    
    combined_output = "\n".join(combined_output_lines)
    return combined_output

# Local formatting function for manual counting
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
    towers_data = {}
    
    for _, row in chunk.iterrows():
        tower_name = row['tower_name']
        activity_name = row['activityName']
        count = int(row['CompletedCount'])
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        towers_data[tower_name].append({
            "activityName": activity_name,
            "completedCount": count
        })
    
    output = ""
    total_activities = 0
    
    for tower_name, activities in sorted(towers_data.items()):
        output += f"Tower: {tower_name}\n"
        output += "activityName            CompletedCount\n"
        activity_dict = {}
        for activity in activities:
            name = activity['activityName']
            count = activity['completedCount']
            activity_dict[name] = activity_dict.get(name, 0) + count
        for name, count in sorted(activity_dict.items()):
            output += f"{name:<30} {count}\n"
            total_activities += count
    
    output += f"Total Completed Activities: {total_activities}"
    return output

def process_data(df, activity_df, location_df, dataset_name):
    completed = df[df['statusName'] == 'Completed']
    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data.")
        return pd.DataFrame(), 0

    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

    if 'qiActivityId' not in completed.columns:
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

    if completed['name'].isna().all():
        logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
        st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
        completed['name'] = 'Unknown'
    else:
        completed['name'] = completed['name'].fillna('Unknown')

    completed['activityName'] = completed['activityName'].fillna('Unknown')

    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        path = []
        current_id = location_id
        max_depth = 10
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
                break
            
            parent_id = parent_child_dict.get(current_id)
            name = name_dict.get(current_id, "Unknown")
            
            if not parent_id:
                if name != "Quality":
                    path.append(name)
                    path.append("Quality")
                else:
                    path.append(name)
                break
            
            path.append(name)
            current_id = parent_id
            depth += 1
        
        if depth >= max_depth:
            logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
        if not path:
            logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
            return "Unknown"
        
        full_path = '/'.join(reversed(path))
        logger.debug(f"Full path for location_id {location_id}: {full_path}")
        return full_path

    completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

    def has_flat_number(full_path):
        parts = full_path.split('/')
        last_part = parts[-1]
        match = re.match(r'^\d+(?:(?:\s*\(LL\))|(?:\s*\(UL\))|(?:\s*LL)|(?:\s*UL))?$', last_part)
        return bool(match)
        
    completed = completed[completed['full_path'].apply(has_flat_number)]
    if completed.empty:
        logger.warning(f"No completed activities with flat numbers found in {dataset_name} data after filtering.")
        return pd.DataFrame(), 0

    def get_tower_name(full_path):
        parts = full_path.split('/')
        if len(parts) < 2:
            return full_path
        
        tower = parts[1]
        if tower == "Tower 4" and len(parts) > 2:
            module = parts[2]
            module_number = module.replace("Module ", "").strip()
            try:
                module_num = int(module_number)
                if 1 <= module_num <= 4:
                    return "Tower 4(B)"
                elif 5 <= module_num <= 8:
                    return "Tower 4(A)"
            except ValueError:
                logger.warning(f"Could not parse module number from {module} in path {full_path}")
        
        return tower

    completed['tower_name'] = completed['full_path'].apply(get_tower_name)

    analysis = completed.groupby(['tower_name', 'activityName'])['qiLocationId'].nunique().reset_index(name='CompletedCount')
    analysis = analysis.sort_values(by=['tower_name', 'activityName'], ascending=True)
    total_completed = analysis['CompletedCount'].sum()

    logger.info(f"Total completed activities for {dataset_name} after processing: {total_completed}")
    return analysis, total_completed

# Main analysis function for Wave City Club Structure
def AnalyzeStatusManually(email=None, password=None):
    start_time = time.time()

    if 'sessionid' not in st.session_state:
        st.error("❌ Please log in first!")
        return

    required_data = [
        'eden_structure',
        'structure_activity_data',
        'structure_location_data'
    ]
    
    for data_key in required_data:
        if data_key not in st.session_state:
            st.error(f"❌ Please fetch required data first! Missing: {data_key}")
            return
        if not isinstance(st.session_state[data_key], pd.DataFrame):
            st.error(f"❌ {data_key} is not a DataFrame! Found type: {type(st.session_state[data_key])}")
            return

    structure_data = st.session_state.eden_structure
    structure_activity = st.session_state.structure_activity_data
    structure_locations = st.session_state.structure_location_data
    
    for df, name in [(structure_data, "Structure")]:
        if 'statusName' not in df.columns:
            st.error(f"❌ statusName column not found in {name} data!")
            return
        if 'qiLocationId' not in df.columns:
            st.error(f"❌ qiLocationId column not found in {name} data!")
            return
        if 'activitySeq' not in df.columns:
            st.error(f"❌ activitySeq column not found in {name} data!")
            return

    for df, name in [(structure_locations, "Structure Location")]:
        if 'qiLocationId' not in df.columns or 'name' not in df.columns:
            st.error(f"❌ qiLocationId or name column not found in {name} data!")
            return

    for df, name in [(structure_activity, "Structure Activity")]:
        if 'activitySeq' not in df.columns or 'activityName' not in df.columns:
            st.error(f"❌ activitySeq or activityName column not found in {name} data!")
            return

    # Process the structure data
    structure_analysis, structure_total = process_data(structure_data, structure_activity, structure_locations, "Structure")

    # Store the structure analysis in session state
    st.session_state.structure_analysis = structure_analysis
    st.session_state.structure_total = structure_total

    st.write("### Wave City Club Structure Quality Analysis (Completed Activities):")
    st.write("**Full Output (Structure):**")
    structure_output = process_manually(structure_analysis, structure_total, "Structure")
    if structure_output:
        st.text(structure_output)
    else:
        st.warning("No structure output generated.")

    end_time = time.time()
    st.write(f"Total execution time: {end_time - start_time:.2f} seconds")

def get_cos_files():
    try:
        # Initialize COS client (assuming initialize_cos_client is defined elsewhere)
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client.")
            return None

        # Update prefix to look for files in the Wave City Club folder
        st.write(f"Attempting to list objects in bucket '{COS_BUCKET}' with prefix 'Wave City Club/'")
        response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix="Wave City Club/")
        if 'Contents' not in response:
            st.error(f"❌ No files found in the 'Wave City Club' folder of bucket '{COS_BUCKET}'.")
            logger.error("No objects found in Wave City Club folder")
            return None

        all_files = [obj['Key'] for obj in response.get('Contents', [])]
        st.write("**All files in Wave City Club folder:**")
        if all_files:
            st.write("\n".join(all_files))
        else:
            st.write("No files found.")
            logger.warning("Wave City Club folder is empty")
            return None

        # Update the regex pattern to match the new file name format
        pattern = re.compile(
            r"Wave City Club/Structure\s*Work\s*Tracker\s*Wave\s*City\s*Club\s*all\s*Block[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
            re.IGNORECASE
        )
        
        # Supported date formats for parsing
        date_formats = ["%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]

        file_info = []
        for obj in response.get('Contents', []):
            key = obj['Key']
            match = pattern.match(key)
            if match:
                date_str = match.group(1).strip('()').strip()
                parsed_date = None
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue
                if parsed_date:
                    file_info.append({'key': key, 'date': parsed_date})
                else:
                    logger.warning(f"Could not parse date in filename: {key}")
                    st.warning(f"Skipping file with unparseable date: {key}")
            else:
                st.write(f"File '{key}' does not match the expected pattern 'Wave City Club/Structure Work Tracker Wave City Club all Block (DD-MM-YYYY).xlsx'")

        if not file_info:
            st.error("❌ No Excel files matched the expected pattern in the 'Wave City Club' folder.")
            logger.error("No files matched the expected pattern")
            return None

        # Find the latest file based on the parsed date
        latest_file = max(file_info, key=lambda x: x['date']) if file_info else None
        if not latest_file:
            st.error("❌ No valid Excel files found for Structure Work Tracker.")
            logger.error("No valid files after date parsing")
            return None

        file_key = latest_file['key']
        st.success(f"Found matching file: {file_key}")
        return file_key
    except Exception as e:
        st.error(f"❌ Error fetching COS files: {str(e)}")
        logger.error(f"Error fetching COS files: {str(e)}")
        return None

if 'cos_df_B1' not in st.session_state:
    st.session_state.cos_df_B1 = None  # For B1 Banket Hall & Finedine
if 'cos_df_B5' not in st.session_state:
    st.session_state.cos_df_B5 = None
if 'cos_df_B6' not in st.session_state:
    st.session_state.cos_df_B6 = None
if 'cos_df_B7' not in st.session_state:
    st.session_state.cos_df_B7 = None
if 'cos_df_B9' not in st.session_state:
    st.session_state.cos_df_B9 = None
if 'cos_df_B8' not in st.session_state:
    st.session_state.cos_df_B8 = None
if 'cos_df_B2_B3' not in st.session_state:
    st.session_state.cos_df_B2_B3 = None  # For B2 & B3
if 'cos_df_B4' not in st.session_state:
    st.session_state.cos_df_B4 = None
if 'cos_df_B11' not in st.session_state:
    st.session_state.cos_df_B11 = None
if 'cos_df_B10' not in st.session_state:
    st.session_state.cos_df_B10 = None

if 'cos_tname_B1' not in st.session_state:
    st.session_state.cos_tname_B1 = None  # For B1 Banket Hall & Finedine
if 'cos_tname_B5' not in st.session_state:
    st.session_state.cos_tname_B5 = None
if 'cos_tname_B6' not in st.session_state:
    st.session_state.cos_tname_B6 = None
if 'cos_tname_B7' not in st.session_state:
    st.session_state.cos_tname_B7 = None
if 'cos_tname_B9' not in st.session_state:
    st.session_state.cos_tname_B9 = None
if 'cos_tname_B8' not in st.session_state:
    st.session_state.cos_tname_B8 = None
if 'cos_tname_B2_B3' not in st.session_state:
    st.session_state.cos_tname_B2_B3 = None  # For B2 & B3
if 'cos_tname_B4' not in st.session_state:
    st.session_state.cos_tname_B4 = None
if 'cos_tname_B11' not in st.session_state:
    st.session_state.cos_tname_B11 = None
if 'cos_tname_B10' not in st.session_state:
    st.session_state.cos_tname_B10 = None

if 'ai_response' not in st.session_state:
    st.session_state.ai_response = {}  # Initialize as empty dictionary

# Process Excel files for Wave City Club blocks with updated sheet names and expected_columns
def process_file(file_stream, filename):
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")

        # Update target sheets to match the exact names in the Excel file
        target_sheets = [
            "B1 Banket Hall & Finedine ",  # Include the trailing space
            "B5", "B6", "B7", "B9", "B8", 
            "B2 & B3",  # Update to match the exact format in the Excel file
            "B4", "B11", "B10"
        ]
        results = []

        # Update expected columns to match the actual 31 columns in the Excel file
        expected_columns = [
            'Block', 'Floor', 'Part', 'Domain', 'Monthly Look Ahead', 'Activity ID', 
            'Activity Name', 'Duration', 'Baseline Start', 'Baseline Finish', 
            'Actual Start', 'Actual Finish', '% Complete', 'Start ()', 'Finish ()', 
            'Actual Start ()', 'Forecast start', 'Forecast End', 'Forecast Today', 
            'Duration.1', 'Balance work', 'Per day', 'Month plan', 'Week 1', 
            'Week 2', 'Week 3', 'Week 4', 'Total For the month', 'Till date', 
            'Average', 'No of days'
        ]

        for sheet_name in target_sheets:
            if sheet_name not in available_sheets:
                st.warning(f"Sheet '{sheet_name}' not found in file: {filename}")
                continue

            file_stream.seek(0)

            try:
                df = pd.read_excel(file_stream, sheet_name=sheet_name, header=1)
                st.write(f"Raw columns in {sheet_name}: {list(df.columns)}")

                # Trim any leading/trailing spaces in column names (e.g., 'Block ' in B11, B10)
                df.columns = [col.strip() for col in df.columns]

                if len(df.columns) != len(expected_columns):
                    st.error(f"Sheet {sheet_name} has {len(df.columns)} columns, but {len(expected_columns)} were expected: {list(df.columns)}")
                    continue

                df.columns = expected_columns

                # Update target_columns to use 'Activity Name' instead of 'Task Name'
                target_columns = ["Activity Name", "Actual Start", "Actual Finish"]
                available_columns = [col for col in target_columns if col in df.columns]

                if len(available_columns) < len(target_columns):
                    missing_cols = [col for col in target_columns if col not in available_columns]
                    st.warning(f"Missing columns in sheet {sheet_name}: {', '.join(missing_cols)}")
                    for col in missing_cols:
                        df[col] = None

                df_original = df.copy()
                df = df[target_columns]
                df = df.dropna(subset=['Activity Name'])  # Update to use 'Activity Name'
                df['Activity Name'] = df['Activity Name'].astype(str).str.strip()

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    st.write(f"Sample of rows with NA or invalid values in Actual Finish for {sheet_name}:")
                    na_rows = df[has_na_mask][['Activity Name', 'Actual Finish']]  # Update to use 'Activity Name'
                    if not na_rows.empty:
                        st.write(na_rows.head(10))
                    else:
                        st.write("No NA or invalid values found in Actual Finish")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                st.write(f"Unique Activity Names in {sheet_name}:")
                unique_tasks = df[['Activity Name']].drop_duplicates()  # Update to use 'Activity Name'
                st.write(unique_tasks)

                results.append((df, sheet_name))

            except Exception as e:
                st.error(f"Error processing sheet {sheet_name}: {str(e)}")
                continue

        if not results:
            st.error(f"No valid sheets ({', '.join(target_sheets)}) found in file: {filename}")
            return [(None, None)]

        return results

    except Exception as e:
        st.error(f"Error loading Excel file: {str(e)}")
        return [(None, None)]


# Function to get access token for WatsonX API
def get_access_token(api_key):
    try:
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {
            "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
            "apikey": api_key
        }
        response = requests.post("https://iam.cloud.ibm.com/identity/token", headers=headers, data=data)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        logger.error(f"Error getting access token: {str(e)}")
        return None

# WatsonX Prompt Generation
def generatePrompt(json_datas):
    try:
        if isinstance(json_datas, pd.DataFrame):
            json_str = json_datas.reset_index().to_json(orient='records', indent=2)
        else:
            json_str = str(json_datas)

        body = {
            "input": f"""
            Read the table data provided below and categorize the activities into the following categories: MEP, Interior Finishing, Structure Work, and ED Civil. Compute the total count of each activity within its respective category and return the results as a JSON array, following the example format provided. For the MEP category, calculate the minimum count between 'UP-First Fix' and 'CP-First Fix' and report it as 'Min. count of UP-First Fix and CP-First Fix'. If an activity is not found in the data, include it with a count of 0. If the Structure Work category has no activities, include it as an empty array. Ensure the counts are accurate, the output is grouped by category, and the JSON structure is valid with no nested or repeated keys.

            Table Data:
            {json_str}

            Categories and Activities:
            - MEP: EL-First Fix, Min. count of UP-First Fix and CP-First Fix, C-Gypsum and POP Punning, EL-Second Fix, No. of Slab cast, Electrical
            - Interior Finishing: Installation of doors, Waterproofing Works, Wall Tiling, Floor Tiling
            - ED Civil: Sewer Line, Storm Line, GSB, WMM, Stamp Concrete, Saucer drain, Kerb Stone
            - Structure Work: (no activities specified)

            Example JSON format needed:
            [
              {{
                "Category": "MEP",
                "Activities": [
                  {{"Activity Name": "EL-First Fix", "Total": 0}},
                  {{"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0}},
                  {{"Activity Name": "C-Gypsum and POP Punning", "Total": 0}},
                  {{"Activity Name": "EL-Second Fix", "Total": 0}},
                  {{"Activity Name": "No. of Slab cast", "Total": 0}},
                  {{"Activity Name": "Electrical", "Total": 0}}
                ]
              }},
              {{
                "Category": "Interior Finishing",
                "Activities": [
                  {{"Activity Name": "Installation of doors", "Total": 0}},
                  {{"Activity Name": "Waterproofing Works", "Total": 0}},
                  {{"Activity Name": "Wall Tiling", "Total": 0}},
                  {{"Activity Name": "Floor Tiling", "Total": 0}}
                ]
              }},
              {{
                "Category": "Structure Work",
                "Activities": []
              }},
              {{
                "Category": "ED Civil",
                "Activities": [
                  {{"Activity Name": "Sewer Line", "Total": 0}},
                  {{"Activity Name": "Storm Line", "Total": 0}},
                  {{"Activity Name": "GSB", "Total": 0}},
                  {{"Activity Name": "WMM", "Total": 0}},
                  {{"Activity Name": "Stamp Concrete", "Total": 0}},
                  {{"Activity Name": "Saucer drain", "Total": 0}},
                  {{"Activity Name": "Kerb Stone", "Total": 0}}
                ]
              }}
            ]

            Return only the JSON array, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
            """,
            "parameters": {
                "decoding_method": "greedy",
                "max_new_tokens": 8100,
                "min_new_tokens": 0,
                "stop_sequences": [";"],
                "repetition_penalty": 1.05,
                "temperature": 0.5
            },
            "model_id": os.getenv("MODEL_ID_1"),
            "project_id": os.getenv("PROJECT_ID_1")
        }
        
        access_token = get_access_token(os.getenv("API_KEY_1"))
        if not access_token:
            logger.error("Failed to obtain access token for WatsonX API")
            return generate_fallback_totals(json_datas)
            
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        logger.info("Sending request to WatsonX API")
        response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=60)
        
        logger.info(f"WatsonX API response status: {response.status_code}")
        logger.debug(f"WatsonX API response text: {response.text[:1000]}...")  # Log first 1000 chars
        
        if response.status_code != 200:
            logger.error(f"WatsonX API call failed: {response.status_code} - {response.text}")
            st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)
            
        response_data = response.json()
        logger.debug(f"WatsonX API response data: {response_data}")
        
        if 'results' not in response_data or not response_data['results']:
            logger.error("WatsonX API response does not contain 'results' key")
            st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        generated_text = response_data['results'][0].get('generated_text', '').strip()
        logger.debug(f"Generated text from WatsonX: {generated_text[:1000]}...")  # Log first 1000 chars
        
        if not generated_text:
            logger.error("WatsonX API returned empty generated text")
            st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        if not (generated_text.startswith('[') and generated_text.endswith(']')):
            start_idx = generated_text.find('[')
            end_idx = generated_text.rfind(']')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                generated_text = generated_text[start_idx:end_idx+1]
                logger.info("Extracted JSON array from response")
            else:
                logger.error(f"Could not extract valid JSON array from response: {generated_text[:1000]}...")
                return generate_fallback_totals(json_datas)
        
        try:
            parsed_json = json.loads(generated_text)
            if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in parsed_json):
                logger.warning("JSON structure doesn't match expected format")
                return generate_fallback_totals(json_datas)
            logger.info("Successfully parsed WatsonX API response")
            return generated_text
        except json.JSONDecodeError as e:
            logger.error(f"WatsonX API returned invalid JSON: {e}")
            st.warning(f"WatsonX API returned invalid JSON. Error: {str(e)}. Using fallback method to calculate totals.")
            error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
            context_start = max(0, error_position - 50)
            context_end = min(len(generated_text), error_position + 50)
            logger.error(f"JSON error context: ...{generated_text[context_start:error_position]}[ERROR HERE]{generated_text[error_position:context_end]}...")
            return generate_fallback_totals(json_datas)
    
    except Exception as e:
        logger.error(f"Error in WatsonX API call: {str(e)}")
        st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
        return generate_fallback_totals(json_datas)

# Fallback Total Calculation
def generate_fallback_totals(count_table):
    try:
        if not isinstance(count_table, pd.DataFrame):
            logger.error("Fallback method received invalid input: not a DataFrame")
            return json.dumps([
                {"Category": "MEP", "Activities": [
                    {"Activity Name": "EL-First Fix", "Total": 0},
                    {"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0},
                    {"Activity Name": "C-Gypsum and POP Punning", "Total": 0},
                    {"Activity Name": "EL-Second Fix", "Total": 0},
                    {"Activity Name": "No. of Slab cast", "Total": 0},
                    {"Activity Name": "Electrical", "Total": 0}
                ]},
                {"Category": "Interior Finishing", "Activities": [
                    {"Activity Name": "Installation of doors", "Total": 0},
                    {"Activity Name": "Waterproofing Works", "Total": 0},
                    {"Activity Name": "Wall Tiling", "Total": 0},
                    {"Activity Name": "Floor Tiling", "Total": 0}
                ]},
                {"Category": "Structure Work", "Activities": []},
                {"Category": "ED Civil", "Activities": [
                    {"Activity Name": "Sewer Line", "Total": 0},
                    {"Activity Name": "Storm Line", "Total": 0},
                    {"Activity Name": "GSB", "Total": 0},
                    {"Activity Name": "WMM", "Total": 0},
                    {"Activity Name": "Stamp Concrete", "Total": 0},
                    {"Activity Name": "Saucer drain", "Total": 0},
                    {"Activity Name": "Kerb Stone", "Total": 0}
                ]}
            ], indent=2)

        categories = {
            "MEP": [
                "EL-First Fix", "Min. count of UP-First Fix and CP-First Fix",
                "C-Gypsum and POP Punning", "EL-Second Fix", "No. of Slab cast", "Electrical"
            ],
            "Interior Finishing": [
                "Installation of doors", "Waterproofing Works", "Wall Tiling", "Floor Tiling"
            ],
            "Structure Work": [],
            "ED Civil": [
                "Sewer Line", "Storm Line", "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone"
            ]
        }

        result = []
        for category, activities in categories.items():
            category_data = {"Category": category, "Activities": []}
            
            if category == "MEP":
                for activity in activities:
                    if activity == "Min. count of UP-First Fix and CP-First Fix":
                        combined_count = count_table.loc["UP-First Fix and CP-First Fix", "Count"] if "UP-First Fix and CP-First Fix" in count_table.index else 0
                        total = combined_count
                    else:
                        total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                    category_data["Activities"].append({
                        "Activity Name": activity,
                        "Total": int(total) if pd.notna(total) else 0
                    })
            elif category == "Structure Work":
                category_data["Activities"] = []
            else:
                for activity in activities:
                    total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                    category_data["Activities"].append({
                        "Activity Name": activity,
                        "Total": int(total) if pd.notna(total) else 0
                    })
            
            result.append(category_data)

        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error in fallback total calculation: {str(e)}")
        st.error(f"Error in fallback total calculation: {str(e)}")
        return json.dumps([
            {"Category": "MEP", "Activities": [
                {"Activity Name": "EL-First Fix", "Total": 0},
                {"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0},
                {"Activity Name": "C-Gypsum and POP Punning", "Total": 0},
                {"Activity Name": "EL-Second Fix", "Total": 0},
                {"Activity Name": "No. of Slab cast", "Total": 0},
                {"Activity Name": "Electrical", "Total": 0}
            ]},
            {"Category": "Interior Finishing", "Activities": [
                {"Activity Name": "Installation of doors", "Total": 0},
                {"Activity Name": "Waterproofing Works", "Total": 0},
                {"Activity Name": "Wall Tiling", "Total": 0},
                {"Activity Name": "Floor Tiling", "Total": 0}
            ]},
            {"Category": "Structure Work", "Activities": []},
            {"Category": "ED Civil", "Activities": [
                {"Activity Name": "Sewer Line", "Total": 0},
                {"Activity Name": "Storm Line", "Total": 0},
                {"Activity Name": "GSB", "Total": 0},
                {"Activity Name": "WMM", "Total": 0},
                {"Activity Name": "Stamp Concrete", "Total": 0},
                {"Activity Name": "Saucer drain", "Total": 0},
                {"Activity Name": "Kerb Stone", "Total": 0}
            ]}
        ], indent=2)

# Extract Totals from AI Data
def getTotal(ai_data):
    try:
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
            
        if not isinstance(ai_data, list):
            logger.error(f"AI data is not a list: {ai_data}")
            return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

        share = []
        for category_data in ai_data:
            if isinstance(category_data, dict) and 'Activities' in category_data:
                for activity in category_data['Activities']:
                    if isinstance(activity, dict) and 'Total' in activity:
                        total = activity['Total']
                        share.append(int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0)
                    else:
                        share.append(0)
            else:
                share.append(0)
        return share
    except Exception as e:
        logger.error(f"Error parsing AI data: {str(e)}")
        st.error(f"Error parsing AI data: {str(e)}")
        return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

# Function to handle activity count display
def display_activity_count():
    specific_activities = [
        "EL-First Fix", "Installation of doors", "Waterproofing Works",
        "C-Gypsum and POP Punning", "Wall Tiling", "Floor Tiling",
        "EL-Second Fix", "No. of Slab cast", "Sewer Line", "Storm Line",
        "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone", "Electrical"
    ]
    all_activities = specific_activities + ["UP-First Fix and CP-First Fix"]

    category_mapping = {
        "EL-First Fix": "MEP",
        "UP-First Fix and CP-First Fix": "MEP",
        "C-Gypsum and POP Punning": "MEP",
        "EL-Second Fix": "MEP",
        "No. of Slab cast": "MEP",
        "Electrical": "MEP",
        "Installation of doors": "Interior Finishing",
        "Waterproofing Works": "Interior Finishing",
        "Wall Tiling": "Interior Finishing",
        "Floor Tiling": "Interior Finishing",
        "Sewer Line": "ED Civil",
        "Storm Line": "ED Civil",
        "GSB": "ED Civil",
        "WMM": "ED Civil",
        "Stamp Concrete": "ED Civil",
        "Saucer drain": "ED Civil",
        "Kerb Stone": "ED Civil"
    }

    count_tables = {}
    if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
        st.session_state.ai_response = {}
        logger.info("Re-initialized st.session_state.ai_response as empty dictionary")

    def process_block_data(block_data, tname):
        if block_data is None or block_data.empty:
            logger.warning(f"No data available for {tname}")
            return tname, None

        block_data = block_data.copy()
        
        st.write(f"Debug - First few rows from {tname}:")
        st.write(block_data.head(3))
        
        st.write(f"Debug - Activity Name matches in {tname}:")
        for activity in specific_activities:
            exact_matches = len(block_data[block_data['Activity Name'] == activity])
            st.write(f"{activity}: {exact_matches} exact matches")
        
        up_matches = len(block_data[block_data['Activity Name'] == "UP-First Fix"])
        cp_matches = len(block_data[block_data['Activity Name'] == "CP-First Fix"])
        st.write(f"UP-First Fix: {up_matches} exact matches")
        st.write(f"CP-First Fix: {cp_matches} exact matches")
        
        count_table = pd.DataFrame({
            'Count_Unfiltered': [0] * len(all_activities),
            'Count_Filtered': [0] * len(all_activities)
        }, index=all_activities)
        
        block_data_filtered = block_data.copy()
        if 'Actual Finish' in block_data.columns:
            block_data['Actual_Finish_Original'] = block_data['Actual Finish'].astype(str)
            block_data['Actual Finish'] = pd.to_datetime(block_data['Actual Finish'], errors='coerce')
            has_na_mask = (
                pd.isna(block_data['Actual Finish']) | 
                (block_data['Actual_Finish_Original'].str.upper() == 'NAT') |
                (block_data['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
            )
            block_data_filtered = block_data[~has_na_mask].copy()
            block_data.drop('Actual_Finish_Original', axis=1, inplace=True)
        
        for activity in specific_activities:
            exact_matches = block_data[block_data['Activity Name'] == activity]
            if len(exact_matches) > 0:
                count_table.loc[activity, 'Count_Unfiltered'] = len(exact_matches)
            else:
                case_insensitive_matches = block_data[block_data['Activity Name'].str.lower() == activity.lower()]
                count_table.loc[activity, 'Count_Unfiltered'] = len(case_insensitive_matches)
            
            exact_matches_filtered = block_data_filtered[block_data_filtered['Activity Name'] == activity]
            if len(exact_matches_filtered) > 0:
                count_table.loc[activity, 'Count_Filtered'] = len(exact_matches_filtered)
            else:
                case_insensitive_matches_filtered = block_data_filtered[block_data_filtered['Activity Name'].str.lower() == activity.lower()]
                count_table.loc[activity, 'Count_Filtered'] = len(case_insensitive_matches_filtered)
        
        up_first_fix_matches = block_data[block_data['Activity Name'].str.lower() == "up-first fix".lower()]
        cp_first_fix_matches = block_data[block_data['Activity Name'].str.lower() == "cp-first fix".lower()]
        up_first_fix_count = len(up_first_fix_matches)
        cp_first_fix_count = len(cp_first_fix_matches)
        count_table.loc["UP-First Fix and CP-First Fix", "Count_Unfiltered"] = up_first_fix_count + cp_first_fix_count
        
        up_first_fix_matches_filtered = block_data_filtered[block_data_filtered['Activity Name'].str.lower() == "up-first fix".lower()]
        cp_first_fix_matches_filtered = block_data_filtered[block_data_filtered['Activity Name'].str.lower() == "cp-first fix".lower()]
        up_first_fix_count_filtered = len(up_first_fix_matches_filtered)
        cp_first_fix_count_filtered = len(cp_first_fix_matches_filtered)
        count_table.loc["UP-First Fix and CP-First Fix", "Count_Filtered"] = up_first_fix_count_filtered + cp_first_fix_count_filtered
        
        count_table['Count_Unfiltered'] = count_table['Count_Unfiltered'].astype(int)
        count_table['Count_Filtered'] = count_table['Count_Filtered'].astype(int)
        
        return tname, count_table

    # Process each block's data
    for block, tname_key in [
        (st.session_state.cos_df_B1, 'cos_tname_B1'),
        (st.session_state.cos_df_B5, 'cos_tname_B5'),
        (st.session_state.cos_df_B6, 'cos_tname_B6'),
        (st.session_state.cos_df_B7, 'cos_tname_B7'),
        (st.session_state.cos_df_B9, 'cos_tname_B9'),
        (st.session_state.cos_df_B8, 'cos_tname_B8'),
        (st.session_state.cos_df_B2_B3, 'cos_tname_B2_B3'),
        (st.session_state.cos_df_B4, 'cos_tname_B4'),
        (st.session_state.cos_df_B11, 'cos_tname_B11'),
        (st.session_state.cos_df_B10, 'cos_tname_B10')
    ]:
        if block is not None:
            tname = st.session_state.get(tname_key)
            tname, count_table = process_block_data(block, tname)
            if count_table is not None:
                count_tables[tname] = count_table

    if not count_tables:
        st.error("No processed COS data available. Please click 'Fetch COS' first.")
        st.stop()

    for tname, count_table in count_tables.items():
        with st.spinner(f"Processing activity counts for {tname}..."):
            try:
                st.write(f"Activity Count for {tname} (Unfiltered vs Filtered):")
                st.write(count_table)
                
                count_table_filtered = count_table[['Count_Filtered']].rename(columns={'Count_Filtered': 'Count'})
                ai_response = generatePrompt(count_table_filtered)
                
                try:
                    ai_data = json.loads(ai_response)
                    if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in ai_data):
                        logger.warning(f"Invalid AI data structure for {tname}: {ai_data}")
                        ai_response = generate_fallback_totals(count_table_filtered)
                        ai_data = json.loads(ai_response)
                    
                    st.session_state.ai_response[tname] = ai_data
                    logger.info(f"Stored AI response for {tname}: {ai_data}")
                    
                    totals_mapping = {}
                    for category_data in ai_data:
                        for activity in category_data['Activities']:
                            totals_mapping[activity['Activity Name']] = activity['Total']
                    
                    display_df = count_table.reset_index()
                    display_df.rename(columns={'index': 'Activity Name'}, inplace=True)
                    
                    display_df['Total'] = display_df['Activity Name'].map(
                        lambda x: totals_mapping.get(x, display_df.loc[display_df['Activity Name'] == x, 'Count_Filtered'].iloc[0])
                    )
                    
                    display_df['Category'] = display_df['Activity Name'].map(lambda x: category_mapping.get(x, "Other"))
                    
                    display_df = display_df.sort_values(['Category', 'Activity Name'])
                    
                    st.write(f"Activity Count with Totals for {tname}:")
                    st.write(display_df[['Activity Name', 'Count_Unfiltered', 'Total', 'Category']])
                    
                    st.write(f"Activity Counts by Category for {tname}:")
                    for category in ['MEP', 'Interior Finishing', 'ED Civil', 'Structure Work']:
                        category_df = display_df[display_df['Category'] == category]
                        if not category_df.empty:
                            st.write(f"**{category}**")
                            st.write(category_df[['Activity Name', 'Count_Filtered', 'Total']])
                    
                except Exception as e:
                    logger.error(f"Error processing WatsonX for {tname}: {str(e)}")
                    st.warning(f"Failed to process AI-generated totals for {tname}. Using fallback method.")
                    
                    ai_response = generate_fallback_totals(count_table_filtered)
                    ai_data = json.loads(ai_response)
                    st.session_state.ai_response[tname] = ai_data
                    logger.info(f"Stored fallback AI response for {tname}: {ai_data}")
                    
                    display_df = count_table.reset_index()
                    display_df.rename(columns={'index': 'Activity Name'}, inplace=True)
                    display_df['Category'] = display_df['Activity Name'].map(lambda x: category_mapping.get(x, "Other"))
                    display_df['Total'] = display_df['Count_Filtered']
                    display_df = display_df.sort_values(['Category', 'Activity Name'])
                    
                    st.write(f"Activity Counts by Category for {tname} (using raw counts):")
                    for category in ['MEP', 'Interior Finishing', 'ED Civil', 'Structure Work']:
                        category_df = display_df[display_df['Category'] == category]
                        if not category_df.empty:
                            st.write(f"**{category}**")
                            st.write(category_df[['Activity Name', 'Count_Filtered', 'Total']])
                
            except Exception as e:
                logger.error(f"Error displaying activity count for {tname}: {str(e)}")
                st.error(f"Error displaying activity count for {tname}: {str(e)}")
                count_table_filtered = count_table[['Count_Filtered']].rename(columns={'Count_Filtered': 'Count'})
                ai_response = generate_fallback_totals(count_table_filtered)
                ai_data = json.loads(ai_response)
                st.session_state.ai_response[tname] = ai_data
                logger.info(f"Stored fallback AI response for {tname} on error: {ai_data}")


# Combined function for Initialize and Fetch Data
async def initialize_and_fetch_data(email, password):
    with st.spinner("Starting initialization and data fetching process..."):
        # Step 1: Login
        if not email or not password:
            st.sidebar.error("Please provide both email and password!")
            return False
        try:
            st.sidebar.write("Logging in...")
            session_id = await login_to_asite(email, password)
            if not session_id:
                st.sidebar.error("Login failed!")
                return False
            st.sidebar.success("Login successful!")
        except Exception as e:
            st.sidebar.error(f"Login failed: {str(e)}")
            return False

        # Step 2: Get Workspace ID
        try:
            st.sidebar.write("Fetching Workspace ID...")
            await GetWorkspaceID()
            st.sidebar.success("Workspace ID fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
            return False

        # Step 3: Get Project IDs
        try:
            st.sidebar.write("Fetching Project IDs...")
            await GetProjectId()
            st.sidebar.success("Project IDs fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
            return False

        # Step 4: Get All Data (Structure only)
        try:
            st.sidebar.write("Fetching All Data...")
            Edenstructure = await GetAllDatas()
            st.session_state.eden_structure = Edenstructure
            st.sidebar.success("All Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
            return False

        # Step 5: Get Activity Data
        try:
            st.sidebar.write("Fetching Activity Data...")
            structure_activity_data = await Get_Activity()
            st.session_state.structure_activity_data = structure_activity_data
            st.sidebar.success("Activity Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
            return False

        # Step 6: Get Location/Module Data
        try:
            st.sidebar.write("Fetching Location/Module Data...")
            structure_location_data = await Get_Location()
            st.session_state.structure_location_data = structure_location_data 
            st.sidebar.success("Location/Module Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
            return False

        # Step 7: Fetch COS Files
        try:
            st.sidebar.write("Fetching COS files from Wave City Club folder...")
            file_key = get_cos_files()
            st.session_state.file_key = file_key
            if file_key:
                st.success(f"Found 1 file in COS storage: {file_key}")
                try:
                    st.write(f"Processing file: {file_key}")
                    cos_client = initialize_cos_client()
                    if not cos_client:
                        st.error("Failed to initialize COS client during file fetch")
                        logger.error("COS client initialization failed during file fetch")
                        return False
                    st.write("Fetching file from COS...")
                    response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
                    file_bytes = io.BytesIO(response['Body'].read())
                    st.write("File fetched successfully. Processing sheets...")
                    results = process_file(file_bytes, file_key)
                    st.write(f"Processing results: {len(results)} sheets processed")
                    for df, sheet_name in results:
                        if df is not None:
                            if sheet_name == "B1 Banket Hall & Finedine ":
                                st.session_state.cos_df_B1 = df
                                st.session_state.cos_tname_B1 = "B1 Banket Hall & Finedine"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B5":
                                st.session_state.cos_df_B5 = df
                                st.session_state.cos_tname_B5 = "B5"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B6":
                                st.session_state.cos_df_B6 = df
                                st.session_state.cos_tname_B6 = "B6"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B7":
                                st.session_state.cos_df_B7 = df
                                st.session_state.cos_tname_B7 = "B7"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B9":
                                st.session_state.cos_df_B9 = df
                                st.session_state.cos_tname_B9 = "B9"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B8":
                                st.session_state.cos_df_B8 = df
                                st.session_state.cos_tname_B8 = "B8"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B2 & B3":
                                st.session_state.cos_df_B2_B3 = df
                                st.session_state.cos_tname_B2_B3 = "B2 & B3"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B4":
                                st.session_state.cos_df_B4 = df
                                st.session_state.cos_tname_B4 = "B4"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B11":
                                st.session_state.cos_df_B11 = df
                                st.session_state.cos_tname_B11 = "B11"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B10":
                                st.session_state.cos_df_B10 = df
                                st.session_state.cos_tname_B10 = "B10"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                        else:
                            st.warning(f"No data processed for {sheet_name} in {file_key}.")
                except Exception as e:
                    st.error(f"Error loading {file_key} from cloud storage: {str(e)}")
                    logger.error(f"Error loading {file_key}: {str(e)}")
                    return False
            else:
                st.warning("No expected Excel files available in the 'Wave City Club' folder of the COS bucket.")
                return False
        except Exception as e:
            st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
            logger.error(f"Failed to fetch COS files: {str(e)}")
            return False

    st.sidebar.success("All steps completed successfully!")
    return True


def generate_consolidated_Checklist_excel(structure_analysis, activity_counts):
    try:
        categories = {
            "Interior Finishing (Civil)": ["Installation of doors", "Waterproofing Works", "Wall Tiling", "Floor Tiling"],
            "MEP": ["EL-First Fix", "Plumbing Works", "C-Gypsum and POP Punning", "EL-Second Fix", "No. of Slab cast", "Electrical"],
            "Structure": [],
            "External Development (Civil)": ["Sewer Line", "Storm Line", "GSB", "WMM", "Stamp Concrete", "Saucer drain", "Kerb Stone"],
            "External Development (MEP)": []
        }

        cos_to_asite_mapping = {
            "EL-First Fix": "Wall Conducting",
            "Installation of doors": ["Door/Window Frame", "Door/Window Shutter"],
            "Plumbing Works": "Plumbing Works",
            "Waterproofing Works": "Waterproofing - Sunken",
            "C-Gypsum and POP Punning": "POP & Gypsum Plaster",
            "Wall Tiling": "Wall Tile",
            "Floor Tiling": "Floor Tiling",
            "EL-Second Fix": "Wiring & Switch Socket",
            "No. of Slab cast": "No. of Slab cast",
            "Sewer Line": "Sewer Line",
            "Storm Line": "Rain Water/Storm",
            "GSB": "Granular Sub-base",
            "WMM": "WMM",
            "Saucer drain": "Saucer drain/Paver block",
            "Kerb Stone": "Kerb Stone",
            "Electrical": "Electrical Cable",
            "Stamp Concrete": "Concreting"
        }

        blocks = [
            "B1 Banket Hall & Finedine", "B5", "B6", "B7", "B9", "B8",
            "B2 & B3", "B4", "B11", "B10"
        ]

        consolidated_rows = []

        for block in blocks:
            block_key = block
            for category, activities in categories.items():
                if not activities and "Structure" not in category:
                    continue

                if activities:
                    for activity in activities:
                        asite_activity = cos_to_asite_mapping.get(activity, activity)
                        if isinstance(asite_activity, list):
                            asite_activities = asite_activity
                        else:
                            asite_activities = [asite_activity]

                        closed_checklist = 0
                        if structure_analysis is not None and not structure_analysis.empty:
                            for asite_act in asite_activities:
                                matching_rows = structure_analysis[
                                    (structure_analysis['tower_name'] == block) &
                                    (structure_analysis['activityName'] == asite_act)
                                ]
                                closed_checklist += matching_rows['CompletedCount'].sum() if not matching_rows.empty else 0

                        completed_flats = 0
                        if block_key in activity_counts:
                            ai_data = activity_counts[block_key]  # ai_data is a list of category dictionaries
                            # Special handling for Plumbing Works
                            if activity == "Plumbing Works":
                                for cat_data in ai_data:
                                    for act in cat_data['Activities']:
                                        if act['Activity Name'] == "UP-First Fix and CP-First Fix":
                                            completed_flats = act['Total']
                                            break
                            else:
                                for cat_data in ai_data:
                                    for act in cat_data['Activities']:
                                        if act['Activity Name'] == activity:
                                            completed_flats = act['Total']
                                            break

                        in_progress = 0
                        open_missing = abs(completed_flats - closed_checklist)

                        display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

                        consolidated_rows.append({
                            "Block": block,
                            "Category": category,
                            "Activity Name": display_activity,
                            "Completed Work*(Count of Flat)": completed_flats,
                            "In progress": in_progress,
                            "Closed checklist": closed_checklist,
                            "Open/Missing check list": open_missing
                        })
                else:
                    consolidated_rows.append({
                        "Block": block,
                        "Category": category,
                        "Activity Name": "",
                        "Completed Work*(Count of Flat)": 0,
                        "In progress": 0,
                        "Closed checklist": 0,
                        "Open/Missing check list": 0
                    })

        df = pd.DataFrame(consolidated_rows)
        if df.empty:
            st.warning("No data available to generate consolidated checklist.")
            return None

        df.sort_values(by=["Block", "Category"], inplace=True)

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Consolidated Checklist")

        # Define formats
        header_format = workbook.add_format({
            'bold': True, 
            'bg_color': '#D3D3D3',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'
        })
        total_format = workbook.add_format({
            'bold': True, 
            'bg_color': '#FFDAB9',
            'border': 1,
            'align': 'center'
        })
        cell_format = workbook.add_format({'border': 1})

        headers = ["Activity Name", "Completed", "In progress", "Closed checklist", "Open/Missing check list"]

        # Group by block first
        grouped_by_block = df.groupby('Block')
        
        current_row = 0
        
        for block, block_group in grouped_by_block:
            # Group by category within each block
            grouped_by_category = block_group.groupby('Category')
            
            # Calculate number of categories to determine layout
            categories = list(grouped_by_category.groups.keys())
            
            # Position categories side by side (2 per row)
            categories_per_row = 2
            col_width = 6  # Width for each category table
            
            for i, (category, cat_group) in enumerate(grouped_by_category):
                # Calculate position
                row_offset = (i // categories_per_row) * 12  # Each table takes about 12 rows
                col_offset = (i % categories_per_row) * col_width
                
                table_start_row = current_row + row_offset
                table_start_col = col_offset
                
                # Write category header
                category_title = f"{block} {category} Checklist Status"
                worksheet.merge_range(
                    table_start_row, table_start_col, 
                    table_start_row, table_start_col + 4, 
                    category_title, header_format
                )
                
                # Write column headers
                for j, header in enumerate(headers):
                    worksheet.write(table_start_row + 1, table_start_col + j, header, header_format)
                
                # Write data rows
                data_start_row = table_start_row + 2
                total_pending = 0
                
                if not cat_group.empty and cat_group["Activity Name"].iloc[0] != "":
                    for idx, (_, row) in enumerate(cat_group.iterrows()):
                        current_data_row = data_start_row + idx
                        worksheet.write(current_data_row, table_start_col + 0, row["Activity Name"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 1, row["Completed Work*(Count of Flat)"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 2, row["In progress"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 3, row["Closed checklist"], cell_format)
                        worksheet.write(current_data_row, table_start_col + 4, row["Open/Missing check list"], cell_format)
                        total_pending += row["Open/Missing check list"]
                    
                    # Add some empty rows if needed (minimum 5 data rows for consistent table size)
                    min_rows = 5
                    actual_rows = len(cat_group)
                    for empty_row in range(actual_rows, min_rows):
                        current_data_row = data_start_row + empty_row
                        for col in range(5):
                            worksheet.write(current_data_row, table_start_col + col, "", cell_format)
                else:
                    # Empty category - fill with blank rows
                    for empty_row in range(5):
                        current_data_row = data_start_row + empty_row
                        for col in range(5):
                            worksheet.write(current_data_row, table_start_col + col, "", cell_format)
                
                # Write total pending row
                total_row = data_start_row + 5
                worksheet.merge_range(
                    total_row, table_start_col, 
                    total_row, table_start_col + 3, 
                    "Total pending check list", total_format
                )
                worksheet.write(total_row, table_start_col + 4, total_pending, total_format)
            
            # Move to next block (leave space between blocks)
            max_categories_in_row = (len(categories) + categories_per_row - 1) // categories_per_row
            current_row += max_categories_in_row * 12 + 3  # 3 rows gap between blocks
        
        # Set column widths
        for col in range(12):  # Adjust for 2 tables side by side
            worksheet.set_column(col, col, 18)

        workbook.close()
        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating consolidated Excel: {str(e)}")
        st.error(f"❌ Error generating Excel file: {str(e)}")
        return None


# Combined function to handle analysis and display
def run_analysis_and_display():
    try:
        st.write("Running status analysis...")
        AnalyzeStatusManually()
        st.success("Status analysis completed successfully!")

        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
            logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

        st.write("Displaying activity counts and generating AI data...")
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        st.write("Checking AI data totals...")
        logger.info(f"st.session_state.ai_response contents: {st.session_state.ai_response}")
        if not st.session_state.ai_response:
            st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
            logger.error("No AI data in st.session_state.ai_response after display_activity_count")
            display_activity_count()
            if not st.session_state.ai_response:
                st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
                logger.error("Failed to regenerate AI data")
                return

        st.write("Generating consolidated checklist Excel file...")
        structure_analysis = st.session_state.get('structure_analysis', None)
        if structure_analysis is None:
            st.error("❌ No structure analysis data available. Please ensure analysis ran successfully.")
            logger.error("No structure_analysis in st.session_state")
            return

        # Fixed Debug statements - Handle the list structure properly
        st.write("Debug - structure_analysis:")
        if hasattr(structure_analysis, 'head'):
            st.write(structure_analysis.head())
        else:
            st.write(f"Structure analysis type: {type(structure_analysis)}")
            st.write(str(structure_analysis)[:500] + "..." if len(str(structure_analysis)) > 500 else str(structure_analysis))
        
        st.write("Debug - activity_counts keys:")
        st.write(list(st.session_state.ai_response.keys()))
        
        st.write("Debug - activity_counts sample:")
        if st.session_state.ai_response:
            first_key = list(st.session_state.ai_response.keys())[0]
            first_value = st.session_state.ai_response[first_key]
            st.write(f"Type of first value: {type(first_value)}")
            if isinstance(first_value, list):
                st.write(f"Sample data (first few items): {first_value[:3] if len(first_value) > 3 else first_value}")
            elif hasattr(first_value, 'head'):
                st.write(first_value.head())
            else:
                st.write(str(first_value)[:500] + "..." if len(str(first_value)) > 500 else str(first_value))

        with st.spinner("Generating Excel file... This may take a moment."):
            excel_file = generate_consolidated_Checklist_excel(structure_analysis, st.session_state.ai_response)
        
        if excel_file:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_WaveCityClub_{timestamp}.xlsx"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.sidebar.download_button(
                    label="📥 Download Checklist Excel",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button",
                    help="Click to download the consolidated checklist in Excel format."
                )
            st.success("Excel file generated successfully! Click the button above to download.")
        else:
            st.error("Failed to generate Excel file. Please check the logs for details.")
            logger.error("Failed to generate Excel file")

    except Exception as e:
        st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
        logger.error(f"Error during analysis, display, or Excel generation: {str(e)}")

# Streamlit UI
st.markdown(
    """
    <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
               color: red; 
               font-size: 48px; 
               text-align: center;'>
        CheckList - Report
    </h1>
    """,
    unsafe_allow_html=True
)

# Initialize and Fetch Data
st.sidebar.title("🔒 Asite Initialization")
email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

if st.sidebar.button("Initialize and Fetch Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        success = loop.run_until_complete(initialize_and_fetch_data(email, password))
        if success:
            st.sidebar.success("Initialization and data fetching completed successfully!")
        else:
            st.sidebar.error("Initialization and data fetching failed!")
    except Exception as e:
        st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
    finally:
        loop.close()

# Analyze and Display
st.sidebar.title("📊 Status Analysis")
if st.sidebar.button("Analyze and Display Activity Counts"):
    with st.spinner("Running analysis and displaying activity counts..."):
        run_analysis_and_display()

